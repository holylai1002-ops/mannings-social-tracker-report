"""
3_v5_period.py - Master Aggregation for Period-Specific Dashboard (Feed output)

Based on 3_v4_period.py with:
  - Output is the final Feed file
  - Key Metrics tabs include Total Followers + Net Followers Growth from API logs
  - API-sourced tabs (Unique Page View, FB Followers, IG Followers)
  - FB Reach Funnel tab
  - Added LinkedIn tabs (Posts Perf, Page Perf, Follower Log) at the end

Input:
    Data	        Source	                                                            Location in code
    FB Posts	    Google Sheet "All FB Posts_New Cat" tab (via GSHEET_TRACKER_ID)
    IG Stories	    Google Sheet "All_IGS" tab (via GSHEET_TRACKER_ID)
    IG Wall Posts	Google Sheet "All_IG_Posts" tab (via GSHEET_TRACKER_ID)
    FB/IG Followers Google Sheet "FB API Log" / "IG API Log" (via GSHEET_KEY_METRICS_ID)
    Comments	    Local FB Comments - {Y}_{M:02d}.xlsx
    LinkedIn Data   Local directory: C:\\Users\\holylai\\Documents\\n8n\\Mannings\\LinkedIn raw

Output:
  Mannings_FB_IG_Dashboard_Feed_{TARGET_YEAR}_{TARGET_MONTH:02d}.xlsx
"""
import os
import re
import json
import logging
from pathlib import Path
import pandas as pd
import datetime

from config import BASE_DIR, GSHEET_TRACKER_ID, GSHEET_COMMENTS_ID, VALID_YEARS, CREDENTIALS_PATH
from utils import (
    get_gspread_client, clean_excel_characters, clean_numeric_col,
    get_actual_col_name, normalize_columns, smart_parse_dates
)

# ==============================================================================
# CONFIGURATION - TARGET PERIOD
# ==============================================================================
TARGET_YEAR = 2026
TARGET_MONTH = 6
TARGET_PERIOD = f"{TARGET_YEAR}-{TARGET_MONTH:02d}"

FEED_OUTPUT = BASE_DIR / f"Mannings_FB_IG_Dashboard_Feed_{TARGET_YEAR}_{TARGET_MONTH:02d}.xlsx"
COMMENTS_XLSX = BASE_DIR / f"FB Comments - {TARGET_YEAR}_{TARGET_MONTH:02d}.xlsx"
IG_API_FILE = BASE_DIR / "ig_api_data.json"

GSHEET_KEY_METRICS_ID = "1f9HLS0HXs2B-a_fxvvcUKuRKz0iMvmEDVivoLFl2vr8"

# LinkedIn Data Configurations
LINKEDIN_RAW_DIR = Path(r"C:\Users\holylai\Documents\n8n\Mannings\LinkedIn raw")
MONTH_MAP = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ==============================================================================
# STRATEGIC LINKEDIN DATE PARSER
# ==============================================================================
def process_linkedin_dates(series, col_name="Date"):
    """
    Intelligently checks variable types (native datetime64, dash-strings, or slash-strings).
    Converts items to integers to drop leading zeros, producing clean m/d/yyyy strings.
    Returns day element keys to support comprehensive date sorting logic.
    """
    results = []
    months = []
    years = []
    days = []
    
    for idx, val in series.items():
        if pd.isna(val) or str(val).strip().lower() in ['nan', 'none', '']:
            results.append("")
            months.append(None)
            years.append(None)
            days.append(None)
            continue
            
        # Case 1: Pandas auto-parsed the column as a datetime64/Timestamp object
        if isinstance(val, (pd.Timestamp, datetime.date)) or hasattr(val, 'month'):
            m, d, y = val.month, val.day, val.year
            results.append(f"{m}/{d}/{y}")
            months.append(m)
            years.append(y)
            days.append(d)
            continue
            
        # Case 2: Raw String Fallback (Regex-based grouping for dashes/slashes)
        val_str = str(val).strip()
        match = re.search(r'(\d+)[/-](\d+)[/-](\d+)', val_str)
        if match:
            try:
                p1, p2, p3 = match.groups()
                if len(p1) == 4:  # ISO Format string: YYYY-MM-DD
                    y, m, d = int(p1), int(p2), int(p3)
                else:  # Custom Format string: MM/DD/YYYY or M/D/YYYY
                    m, d, y = int(p1), int(p2), int(p3)
                    
                results.append(f"{m}/{d}/{y}")
                months.append(m)
                years.append(y)
                days.append(d)
            except:
                results.append("")
                months.append(None)
                years.append(None)
                days.append(None)
        else:
            results.append("")
            months.append(None)
            years.append(None)
            days.append(None)
        
    return (
        pd.Series(results, index=series.index),
        pd.Series(months, index=series.index),
        pd.Series(years, index=series.index),
        pd.Series(days, index=series.index)
    )
    
# ==============================================================================
# LINKEDIN SMART SHEET INPUT MANAGEMENT ENGINE
# ==============================================================================
def smart_read_linkedin_excel(file_path, sheet_keyword=None):
    """
    Locates the true header block by checking for key identifiers on multiple rows,
    skipping any descriptive metadata text rows exported from LinkedIn.
    """
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = next((s for s in xl.sheet_names if sheet_keyword.lower() in s.lower()), xl.sheet_names[0]) if sheet_keyword else xl.sheet_names[0]
        df = xl.parse(sheet_name, header=None)
    except Exception:
        try:
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
        except Exception as e:
            logging.error(f"  Failed parsing workbook object structural keys for {file_path.name}: {e}")
            return pd.DataFrame()

    header_row_idx = 0
    for i, row in df.iterrows():
        row_strs = [str(x).lower().strip() for x in row.dropna().values]
        if len(row_strs) > 3 and any(s in ['date', 'created date', 'post title', 'impressions', 'followers'] for s in row_strs):
            header_row_idx = i
            break

    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx+1:].reset_index(drop=True)
    df.columns = df.columns.str.strip()
    return df

# ==============================================================================
# PERIOD FILTER
# ==============================================================================
def filter_to_target_period(df, date_col='Publish time', dayfirst_guess=False):
    df.columns = df.columns.str.strip()
    df = df.copy()

    # --- Preferred: filter by "Month No." column if available ---
    month_no_col = None
    for candidate in ['Month No.', 'Month No', 'month no.', 'month no', 'Month']:
        if candidate in df.columns:
            month_no_col = candidate
            break

    if month_no_col:
        month_vals = pd.to_numeric(df[month_no_col], errors='coerce')
        mask = month_vals == TARGET_MONTH

        year_col = None
        for candidate in ['Year', 'year']:
            if candidate in df.columns:
                year_col = candidate
                break
        if year_col:
            year_vals = pd.to_numeric(df[year_col], errors='coerce')
            mask = mask & (year_vals == TARGET_YEAR)

        df = df[mask].copy()
        df['_Period'] = TARGET_PERIOD
        df['_Year'] = TARGET_YEAR
        df['_Month'] = TARGET_MONTH
        if 'Day' in df.columns:
            df['_Day'] = pd.to_numeric(df['Day'], errors='coerce').fillna(0).astype(int)
        else:
            df['_Day'] = 0
        logging.info(f"  Filtered by '{month_no_col}'={TARGET_MONTH}"
                     f"{f' & {year_col}={TARGET_YEAR}' if year_col else ''}"
                     f" -> {len(df)} rows match {TARGET_PERIOD}")
        return df

    # --- Fallback: date parsing ---
    if date_col not in df.columns:
        logging.warning(f"Column '{date_col}' not found and no 'Month No.' column! Available: {list(df.columns)[:10]}")
        df['_Period'] = TARGET_PERIOD
        df['_Year'] = TARGET_YEAR
        df['_Month'] = TARGET_MONTH
        df['_Day'] = 0
        return df

    time_str = df[date_col].astype(str).str.strip()

    if dayfirst_guess:
        parsed = pd.to_datetime(time_str, dayfirst=True, errors='coerce')
        parsed = parsed.fillna(pd.to_datetime(time_str, format='%m/%d/%Y %H:%M', errors='coerce'))
    else:
        parsed = pd.to_datetime(time_str, format='%m/%d/%Y %H:%M', errors='coerce')
        parsed = parsed.fillna(pd.to_datetime(time_str, errors='coerce'))

    parsed_ok = parsed.notna().sum()
    sample = time_str.iloc[0] if len(time_str) > 0 else 'N/A'
    logging.info(f"  Date parse '{date_col}' (dayfirst={dayfirst_guess}): "
                 f"{len(time_str)} rows, {parsed_ok} parsed, sample='{sample}'")

    mask = (parsed.dt.year == TARGET_YEAR) & (parsed.dt.month == TARGET_MONTH)
    df = df[mask].copy()
    df['_Period'] = TARGET_PERIOD
    df['_Year'] = parsed[mask].dt.year.astype(int).values
    df['_Month'] = parsed[mask].dt.month.astype(int).values
    df['_Day'] = parsed[mask].dt.day.astype(int).values
    logging.info(f"  -> {len(df)} rows match {TARGET_PERIOD}")
    return df


# ==============================================================================
# FB KEY METRICS
# ==============================================================================
def generate_fb_key_metrics(df_fb, target_period):
    """Build FB Key Metrics for the target period. Metrics as rows, period as column."""
    logging.info("Generating FB Key Metrics ...")

    if df_fb is None or df_fb.empty:
        logging.warning("  No FB data for Key Metrics")
        return pd.DataFrame()

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    def safe_count_val(df, col, val):
        if col in df.columns:
            return (df[col].astype(str).str.strip() == val).sum()
        return 0

    feeds_posted = len(df_fb)
    total_dark_posts = safe_count_val(df_fb, 'Dark Post', 'Y')
    total_organic = safe_count_val(df_fb, 'Organic', 'Organic')
    total_paid = safe_count_val(df_fb, 'Paid', 'Paid')
    organic_pct = total_organic / feeds_posted if feeds_posted > 0 else 0
    paid_pct = total_paid / feeds_posted if feeds_posted > 0 else 0

    video_views_col = '3-second video views'
    no_video_posts = 0
    if video_views_col in df_fb.columns:
        col_raw = df_fb[video_views_col].astype(str).str.strip()
        has_content = col_raw.str.len() > 0
        vv = pd.to_numeric(col_raw, errors='coerce')
        no_video_posts = int((has_content & vv.notna()).sum())

    total_interactions = safe_sum(df_fb, 'Interactions')
    total_reactions = safe_sum(df_fb, 'Reactions')
    total_comments = safe_sum(df_fb, 'Comments')
    total_shares = safe_sum(df_fb, 'Shares')
    total_video_views = safe_sum(df_fb, '3-second video views')
    total_organic_reach = safe_sum(df_fb, 'Organic reach')
    total_paid_reach = safe_sum(df_fb, 'Paid reach')

    avg_organic_reach = total_organic_reach / no_video_posts if no_video_posts > 0 else 0
    avg_paid_reach = total_paid_reach / no_video_posts if no_video_posts > 0 else 0
    avg_interaction = total_interactions / no_video_posts if no_video_posts > 0 else 0
    avg_reactions = total_reactions / no_video_posts if no_video_posts > 0 else 0
    avg_comments = total_comments / no_video_posts if no_video_posts > 0 else 0
    avg_shares = total_shares / no_video_posts if no_video_posts > 0 else 0
    avg_video_views = total_video_views / no_video_posts if no_video_posts > 0 else 0

    rows = {
        'Feeds Posted': feeds_posted,
        'Total Dark Posts': total_dark_posts,
        'Total Organic': total_organic,
        'Organic %': organic_pct,
        'Total Paid': total_paid,
        'Paid %': paid_pct,
        'No. of Video Post': no_video_posts,
        'Total Interactions': int(total_interactions),
        'Total Reactions': int(total_reactions),
        'Total Comments': int(total_comments),
        'Total Shares': int(total_shares),
        'Total Video Views': int(total_video_views),
        'Average Organic Reach': round(avg_organic_reach, 1),
        'Average Paid Reach': round(avg_paid_reach, 1),
        'Average Interaction': round(avg_interaction, 1),
        'Average Reactions': round(avg_reactions, 1),
        'Average Comments': round(avg_comments, 1),
        'Average Shares': round(avg_shares, 1),
        'Average Video Views': round(avg_video_views, 1),
    }

    metrics_list = [{'Metric': k, target_period: v} for k, v in rows.items()]
    return pd.DataFrame(metrics_list)


# ==============================================================================
# FB KEY METRIC (CM) - Community Management version
# ==============================================================================
def generate_fb_key_metrics_cm(df_fb, target_period):
    """Build FB Key Metric (CM). Excludes Ecommerce/GNC pillars.
    Averages use No. of Wall Post as denominator."""
    logging.info("Generating FB Key Metric (CM) ...")

    if df_fb is None or df_fb.empty:
        logging.warning("  No FB data for Key Metric (CM)")
        return pd.DataFrame()

    if 'Pillar' in df_fb.columns:
        mask = ~df_fb['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_fb = df_fb[mask].copy()
        logging.info(f"  CM filter (excl Ecommerce/GNC): {len(df_fb)} rows")

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    def safe_count_val(df, col, val):
        if col in df.columns:
            return (df[col].astype(str).str.strip() == val).sum()
        return 0

    feeds_posted = len(df_fb)
    no_dark_post = safe_count_val(df_fb, 'Dark Post', 'Y')
    no_wall_post = feeds_posted - no_dark_post
    no_organic = safe_count_val(df_fb, 'Organic', 'Organic')
    no_paid = safe_count_val(df_fb, 'Paid', 'Paid')
    organic_pct = no_organic / feeds_posted if feeds_posted > 0 else 0
    paid_pct = no_paid / feeds_posted if feeds_posted > 0 else 0

    total_interactions = safe_sum(df_fb, 'Interactions')
    total_reactions = safe_sum(df_fb, 'Reactions')
    total_comments = safe_sum(df_fb, 'Comments')
    total_shares = safe_sum(df_fb, 'Shares')
    total_organic_reach = safe_sum(df_fb, 'Organic reach')
    total_paid_reach = safe_sum(df_fb, 'Paid reach')

    d = no_wall_post
    avg_organic_reach = total_organic_reach / d if d > 0 else 0
    avg_paid_reach = total_paid_reach / d if d > 0 else 0
    avg_interaction = total_interactions / d if d > 0 else 0
    avg_reactions = total_reactions / d if d > 0 else 0
    avg_comments = total_comments / d if d > 0 else 0
    avg_shares = total_shares / d if d > 0 else 0

    rows = {
        'Feeds Posted': feeds_posted,
        'No. of Wall Post': no_wall_post,
        'No. of Dark Post': no_dark_post,
        'No. of Organic Post': no_organic,
        'Organic Post %': round(organic_pct, 4),
        'No. of Paid Post': no_paid,
        'Paid Post %': round(paid_pct, 4),
        'Total Interaction': int(total_interactions),
        'Total Reactions': int(total_reactions),
        'Total Comments': int(total_comments),
        'Total Shares': int(total_shares),
        'Average Organic Reach': round(avg_organic_reach, 1),
        'Average Paid Reach': round(avg_paid_reach, 1),
        'Average Interaction': round(avg_interaction, 1),
        'Average Reactions': round(avg_reactions, 1),
        'Average Comments': round(avg_comments, 1),
        'Average Shares': round(avg_shares, 1),
    }

    metrics_list = [{'Metric': k, target_period: v} for k, v in rows.items()]
    return pd.DataFrame(metrics_list)


# ==============================================================================
# IG KEY METRICS (updated: new order, removed Followers/Engagement Rate)
# ==============================================================================
def generate_ig_key_metrics(df_ig_post, df_ig_story):
    """Build IG Key Metrics for the target period only."""
    logging.info("Generating IG Key Metrics ...")

    if df_ig_post is None or df_ig_post.empty:
        logging.warning("  No IG Post data for Key Metrics")
        return pd.DataFrame()

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    total_posts = len(df_ig_post)
    total_reach = safe_sum(df_ig_post, 'Total Post Reach')
    total_likes = safe_sum(df_ig_post, 'Likes')
    total_comments = safe_sum(df_ig_post, 'Comments')
    total_shares = safe_sum(df_ig_post, 'Shares')
    total_saves = safe_sum(df_ig_post, 'Saves')
    total_interaction = total_likes + total_comments + total_shares + total_saves

    total_stories = 0
    total_story_reach = 0
    total_story_shares = 0
    total_story_clicks = 0
    if df_ig_story is not None and not df_ig_story.empty:
        total_stories = len(df_ig_story)
        total_story_reach = safe_sum(df_ig_story, 'Total Reach')
        total_story_shares = safe_sum(df_ig_story, 'Shares')
        total_story_clicks = safe_sum(df_ig_story, 'Link clicks')

    rows = {
        'Feeds Posted': total_posts,
        'Total Post Likes': int(total_likes),
        'Total Post Comments': int(total_comments),
        'Total Post Shares': int(total_shares),
        'Total Post Saves': int(total_saves),
        'Total Post Interaction': int(total_interaction),
        'Total Post Reach': int(total_reach),
        'Average Post Likes': round(total_likes / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Comments': round(total_comments / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Shares': round(total_shares / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Saves': round(total_saves / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Interaction': round(total_interaction / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Reach': round(total_reach / total_posts, 1) if total_posts > 0 else 0,
        'Stories Posted': total_stories,
        'Total Stories Link Clicks': int(total_story_clicks),
        'Total Stories Shares': int(total_story_shares),
        'Total Stories Reach': int(total_story_reach),
        'Average Stories Link Clicks': round(total_story_clicks / total_stories, 1) if total_stories > 0 else None,
        'Average Stories Shares': round(total_story_shares / total_stories, 1) if total_stories > 0 else None,
        'Average Stories Reach': round(total_story_reach / total_stories, 1) if total_stories > 0 else None,
    }

    metrics_list = [{'Metric': k, TARGET_PERIOD: v} for k, v in rows.items()]
    return pd.DataFrame(metrics_list)


# ==============================================================================
# IG KEY METRIC (CM) - Community Management version
# ==============================================================================
def generate_ig_key_metrics_cm(df_ig_post, df_ig_story):
    """Build IG Key Metric (CM). Excludes Ecommerce/GNC pillars."""
    logging.info("Generating IG Key Metric (CM) ...")

    if df_ig_post is None or df_ig_post.empty:
        logging.warning("  No IG Post data for Key Metric (CM)")
        return pd.DataFrame()

    if 'Pillar' in df_ig_post.columns:
        mask = ~df_ig_post['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_ig_post = df_ig_post[mask].copy()
        logging.info(f"  CM filter IG posts (excl Ecommerce/GNC): {len(df_ig_post)} rows")
    if df_ig_story is not None and not df_ig_story.empty and 'Pillar' in df_ig_story.columns:
        mask = ~df_ig_story['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_ig_story = df_ig_story[mask].copy()
        logging.info(f"  CM filter IG stories (excl Ecommerce/GNC): {len(df_ig_story)} rows")

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    total_posts = len(df_ig_post)
    total_reach = safe_sum(df_ig_post, 'Total Post Reach')
    total_likes = safe_sum(df_ig_post, 'Likes')
    total_comments = safe_sum(df_ig_post, 'Comments')
    total_shares = safe_sum(df_ig_post, 'Shares')
    total_saves = safe_sum(df_ig_post, 'Saves')
    total_interaction = total_likes + total_comments + total_shares + total_saves

    total_stories = 0
    total_story_reach = 0
    total_story_shares = 0
    total_story_clicks = 0
    if df_ig_story is not None and not df_ig_story.empty:
        total_stories = len(df_ig_story)
        total_story_reach = safe_sum(df_ig_story, 'Total Reach')
        total_story_shares = safe_sum(df_ig_story, 'Shares')
        total_story_clicks = safe_sum(df_ig_story, 'Link clicks')

    rows = {
        'Feeds Posted': total_posts,
        'Total Post Likes': int(total_likes),
        'Total Post Comments': int(total_comments),
        'Total Post Shares': int(total_shares),
        'Total Post Saves': int(total_saves),
        'Total Post Interaction': int(total_interaction),
        'Total Post Reach': int(total_reach),
        'Average Post Likes': round(total_likes / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Comments': round(total_comments / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Shares': round(total_shares / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Saves': round(total_saves / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Interaction': round(total_interaction / total_posts, 1) if total_posts > 0 else 0,
        'Average Post Reach': round(total_reach / total_posts, 1) if total_posts > 0 else 0,
        'Stories Posted': total_stories,
        'Total Stories Link Clicks': int(total_story_clicks),
        'Total Stories Shares': int(total_story_shares),
        'Total Stories Reach': int(total_story_reach),
        'Average Stories Link Clicks': round(total_story_clicks / total_stories, 1) if total_stories > 0 else None,
        'Average Stories Shares': round(total_story_shares / total_stories, 1) if total_stories > 0 else None,
        'Average Stories Reach': round(total_story_reach / total_stories, 1) if total_stories > 0 else None,
    }

    metrics_list = [{'Metric': k, TARGET_PERIOD: v} for k, v in rows.items()]
    return pd.DataFrame(metrics_list)


# ==============================================================================
# BULLETPROOF EXCEL CLEANER
# ==============================================================================
_OPENPYXL_ILLEGAL_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff\ufffe\uffff]'
)

def strip_illegal_for_excel(val):
    if val is None:
        return ""
    if isinstance(val, (int, float, bool)):
        return val
    s = str(val)
    s = _OPENPYXL_ILLEGAL_RE.sub('', s)
    s = re.sub(r'[\x7f-\x9f\u200b-\u200f\u2028-\u202f\ufeff]', '', s)
    return s

def deep_clean_df(df):
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].apply(strip_illegal_for_excel)
    return df


_DATE_COL_NAMES = {'Publish time', 'Post Date', 'Comment Date', 'Date'}

def normalize_date_columns(df):
    """Normalize all date-like columns to DD/MM/YYYY HH:MM format."""
    if df is None or df.empty:
        return df
    df = df.copy()
    for col in df.columns:
        col_lower = str(col).lower().strip()
        is_date_col = (col in _DATE_COL_NAMES or
                       'date' in col_lower or
                       'publish time' in col_lower or
                       'post date' in col_lower)
        if not is_date_col:
            continue
        vals = df[col].astype(str).str.strip()
        non_empty = vals[(vals != '') & (vals != 'nan') & (vals != 'None')]
        if len(non_empty) == 0:
            continue
        parsed = smart_parse_dates(vals)
        has_time = non_empty.str.contains(':').any()
        if has_time:
            df[col] = parsed.dt.strftime('%d/%m/%Y %H:%M')
        else:
            df[col] = parsed.dt.strftime('%d/%m/%Y')
    return df


# ==============================================================================
# FOLLOWERS FROM API LOG (for Key Metrics tabs)
# ==============================================================================
def fetch_followers_for_period(worksheet_name):
    """Fetch follower total + net growth for TARGET_PERIOD from API Log."""
    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
        ws = sh.worksheet(worksheet_name)
        df = pd.DataFrame(ws.get_all_records())
        if df.empty or "Date" not in df.columns:
            return {"total": 0, "net": 0}

        df["_dt"] = pd.to_datetime(df["Date"].astype(str).str.strip(),
                                   format="%d/%m/%Y", errors="coerce")
        df = df.dropna(subset=["_dt"]).sort_values("_dt")

        if "Total Followers" in df.columns:
            df["Total Followers"] = pd.to_numeric(
                df["Total Followers"], errors="coerce"
            ).ffill().bfill().astype(int)
        if "Followers Net" in df.columns:
            df["Followers Net"] = pd.to_numeric(
                df["Followers Net"], errors="coerce"
            ).fillna(0).astype(int)

        mask = (df["_dt"].dt.year == TARGET_YEAR) & (df["_dt"].dt.month == TARGET_MONTH)
        group = df[mask]
        if group.empty:
            return {"total": 0, "net": 0}

        total = int(group.iloc[0]["Total Followers"]) if "Total Followers" in df.columns else 0
        net = int(group["Followers Net"].sum()) if "Followers Net" in df.columns else 0
        logging.info(f"  {worksheet_name} {TARGET_PERIOD}: total={total}, net={net}")
        return {"total": total, "net": net}
    except Exception as e:
        logging.error(f"Failed to fetch followers from {worksheet_name}: {e}")
        return {"total": 0, "net": 0}


def add_followers_to_key_metrics(df_metrics, followers_data):
    """Insert Total Followers + Net Followers Growth rows at the top."""
    if not followers_data or df_metrics.empty:
        return df_metrics

    new_rows = pd.DataFrame([
        {"Metric": "Total Followers", TARGET_PERIOD: followers_data["total"]},
        {"Metric": "Net Followers Growth", TARGET_PERIOD: followers_data["net"]},
    ])
    return pd.concat([new_rows, df_metrics], ignore_index=True)


# ==============================================================================
# API-SOURCED TABS + REACH FUNNEL
# ==============================================================================
def build_api_tabs_period():
    """Build API-sourced tabs filtered to target period."""
    logging.info("Building API tabs (period-filtered) ...")
    result = {}

    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
    except Exception as e:
        logging.error(f"Failed to connect for API logs: {e}")
        return result

    # ── Unique Page View + FB Followers (from FB API Log) ──
    df_fb = pd.DataFrame()
    try:
        ws = sh.worksheet("FB API Log")
        df_fb = pd.DataFrame(ws.get_all_records())
        if not df_fb.empty and "Date" in df_fb.columns:
            dt = pd.to_datetime(df_fb["Date"].astype(str).str.strip(),
                                format="%d/%m/%Y", errors="coerce")
            mask = (dt.dt.year == TARGET_YEAR) & (dt.dt.month == TARGET_MONTH)
            df_period = df_fb[mask].copy()

            if not df_period.empty:
                df_view = df_period[["Date", "Unique Page View"]].copy()
                df_view["Unique Page View"] = pd.to_numeric(
                    df_view["Unique Page View"], errors="coerce"
                ).fillna(0).astype(int)
                result["Unique Page View"] = df_view

                cols = ["Date", "Total Followers", "Followers Gain", "Followers Loss", "Followers Net"]
                available = [c for c in cols if c in df_period.columns]
                if available:
                    df_followers = df_period[available].copy()
                    for c in ["Total Followers", "Followers Gain", "Followers Loss", "Followers Net"]:
                        if c in df_followers.columns:
                            df_followers[c] = pd.to_numeric(
                                df_followers[c], errors="coerce"
                            ).fillna(0).astype(int)
                    result["FB Followers"] = df_followers

                logging.info(f"  Unique Page View: {len(df_view)} rows, FB Followers: {len(df_followers)} rows")
    except Exception as e:
        logging.error(f"  FB API Log error: {e}")

    # ── IG Followers (from IG API Log, with forward-fill) ──
    try:
        ws_ig = sh.worksheet("IG API Log")
        df_ig = pd.DataFrame(ws_ig.get_all_records())
        if not df_ig.empty and "Date" in df_ig.columns:
            df_ig["Date"] = df_ig["Date"].astype(str).str.strip()
            df_ig["Followers Net"] = pd.to_numeric(
                df_ig.get("Followers Net", 0), errors="coerce"
            ).fillna(0).astype(int)

            dates = pd.to_datetime(df_ig["Date"], format="%d/%m/%Y", errors="coerce")
            df_ig["_dt"] = dates

            if len(df_ig) > 0:
                min_dt = df_ig["_dt"].min()
                max_dt = df_ig["_dt"].max()
                full_range = pd.date_range(min_dt, max_dt, freq="D")
                df_daily = pd.DataFrame({"_dt": full_range})
                df_daily = df_daily.merge(df_ig, on="_dt", how="left")

                if "Total Followers" in df_daily.columns:
                    tf = pd.to_numeric(df_daily["Total Followers"], errors="coerce")
                    df_daily["Total Followers"] = tf.ffill().bfill().astype(int)

                df_daily["Followers Net"] = df_daily["Followers Net"].fillna(0).astype(int)

                mask_empty = df_daily["Followers Net"] == 0
                if mask_empty.any() and "Total Followers" in df_daily.columns:
                    diffs = df_daily["Total Followers"].diff().fillna(0).astype(int)
                    df_daily.loc[mask_empty, "Followers Net"] = diffs[mask_empty]

                # Filter to target period
                mask = (df_daily["_dt"].dt.year == TARGET_YEAR) & (df_daily["_dt"].dt.month == TARGET_MONTH)
                df_daily = df_daily[mask]

                df_out = pd.DataFrame({
                    "Date": df_daily["_dt"].dt.strftime("%d/%m/%Y"),
                    "Total Followers": df_daily["Total Followers"].astype(int),
                    "Followers Net": df_daily["Followers Net"].astype(int),
                })
                result["IG Followers"] = df_out
                logging.info(f"  IG Followers: {len(df_out)} rows")
    except Exception as e:
        logging.error(f"  IG API Log error: {e}")

    return result


def build_reach_funnel_tab_period():
    """Build FB Reach Funnel tab filtered to target period."""
    logging.info("Building Reach Funnel tab (period-filtered) ...")
    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
        ws = sh.worksheet("All FB Posts_New Cat")
        records = ws.get_all_records()
        df = pd.DataFrame(records)
    except Exception as e:
        logging.error(f"Failed to read All FB Posts_New Cat: {e}")
        return None

    if df.empty:
        return None

    col_map = {}
    for c in df.columns:
        stripped = c.strip()
        if stripped in ("Publish time", "Month No.", "Day", "Organic reach", "Paid reach"):
            col_map[c] = stripped

    mn_col = next((c for c, v in col_map.items() if v == "Month No."), None)
    day_col = next((c for c, v in col_map.items() if v == "Day"), None)
    pt_col = next((c for c, v in col_map.items() if v == "Publish time"), None)

    df_out = pd.DataFrame(index=df.index)
    df_out["Year"] = TARGET_YEAR
    if mn_col:
        df_out["Month No."] = pd.to_numeric(df[mn_col], errors="coerce").astype("Int64")
    if day_col:
        df_out["Day"] = pd.to_numeric(df[day_col], errors="coerce").astype("Int64")
    if pt_col:
        df_out["Publish time"] = df[pt_col].astype(str).str.strip()
    else:
        df_out["Publish time"] = ""

    for label in ("Organic reach", "Paid reach"):
        src_col = next((c for c, v in col_map.items() if v == label), None)
        if src_col:
            df_out[label] = pd.to_numeric(df[src_col].astype(str).str.replace(",", ""), errors="coerce").fillna(0).astype(int)
        else:
            df_out[label] = 0

    df_out = df_out.dropna(subset=["Year", "Month No."])
    df_out = df_out[df_out["Month No."] == TARGET_MONTH].reset_index(drop=True)
    logging.info(f"  Reach Funnel: {len(df_out)} rows for {TARGET_PERIOD}")
    return df_out


# ==============================================================================
# LINKEDIN PERFORMANCE CORE GENERATORS (WITH HYBRID DATE PARSER)
# ==============================================================================
# ==============================================================================
# LINKEDIN PERFORMANCE HARDENED TAB ENGINE
# ==============================================================================
def build_linkedin_posts_perf():
    logging.info("Building LinkedIn Posts Perf tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()

    files = list(LINKEDIN_RAW_DIR.glob("*manningshk_content*"))
    all_dfs = []
    for f in files:
        try:
            df = smart_read_linkedin_excel(f, sheet_keyword='all posts') if f.suffix in ['.xlsx', '.xls'] else pd.read_csv(f, header=None)
            if f.suffix not in ['.xlsx', '.xls']:
                header_row_idx = 0
                for i, row in df.iterrows():
                    row_strs = [str(x).lower().strip() for x in row.dropna().values]
                    if len(row_strs) > 3 and any(s in ['date', 'created date', 'post title', 'impressions'] for s in row_strs):
                        header_row_idx = i
                        break
                df.columns = df.iloc[header_row_idx]
                df = df.iloc[header_row_idx+1:].reset_index(drop=True)
                df.columns = df.columns.str.strip()
            
            if not df.empty: all_dfs.append(df)
        except Exception as e: logging.error(f"  Error reading file {f.name}: {e}")
    if not all_dfs: return pd.DataFrame()

    df_all = pd.concat(all_dfs, ignore_index=True).drop_duplicates()
    date_col = next((c for c in df_all.columns if str(c).lower().strip() in ['created date', 'date']), None)
    if not date_col: return pd.DataFrame()

    # Dynamic Parser + Execution Sorting keys
    fmt_dates, months, years, days = process_linkedin_dates(df_all[date_col], col_name=f"Content Posts: {date_col}")
    df_all['_fmt_date'], df_all['_m'], df_all['_y'], df_all['_d'] = fmt_dates, months, years, days
    df_filtered = df_all[(df_all['_y'] == TARGET_YEAR) & (df_all['_m'] == TARGET_MONTH)].copy()
    if df_filtered.empty: return pd.DataFrame()

    # Chronological Sort Ascending
    df_filtered = df_filtered.sort_values(by=['_y', '_m', '_d'], ascending=True).copy()

    def get_col_or_zero(df, keywords):
        c = next((col for col in df.columns if any(kw in str(col).lower() for kw in keywords)), None)
        return pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int) if c else pd.Series(0, index=df.index)

    df_out = pd.DataFrame()
    df_out['Month'] = df_filtered['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
    df_out['Post date'] = df_filtered['_fmt_date']
    df_out['Post Link'] = df_filtered[next((c for c in df_filtered.columns if any(k in str(c).lower() for k in ['link', 'url'])), df_filtered.columns[0])].astype(str).str.strip()
    df_out['Post message'] = df_filtered[next((c for c in df_filtered.columns if any(kw in str(c).lower() for kw in ['title', 'message', 'description'])), df_filtered.columns[0])].astype(str).str.strip()
    df_out['Impressions'] = get_col_or_zero(df_filtered, ['impressions'])

    ct_col = next((c for c in df_filtered.columns if 'content type' in str(c).lower()), None)
    v_col = next((c for c in df_filtered.columns if 'views' in str(c).lower()), None)
    df_out['Video Views'] = df_filtered.apply(lambda r: pd.to_numeric(r[v_col], errors='coerce') if ct_col and v_col and 'video' in str(r[ct_col]).lower() else 0, axis=1).fillna(0).astype(int)

    df_out['Clicks'] = get_col_or_zero(df_filtered, ['clicks'])
    df_out['Likes'] = get_col_or_zero(df_filtered, ['likes', 'reactions'])
    df_out['Comments'] = get_col_or_zero(df_filtered, ['comments'])
    df_out['Reposts'] = get_col_or_zero(df_filtered, ['reposts', 'shares'])
    df_out['Follows'] = get_col_or_zero(df_filtered, ['follows'])
    df_out['Total Interactions'] = df_out['Likes'] + df_out['Comments'] + df_out['Reposts'] + df_out['Follows']
    return df_out


def build_linkedin_page_perf():
    logging.info("Building LinkedIn Page Perf tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()

    df_metrics = pd.DataFrame()
    for f in LINKEDIN_RAW_DIR.glob("*manningshk_content*"):
        if f.suffix in ['.xlsx', '.xls']:
            s = smart_read_linkedin_excel(f, sheet_keyword='metrics')
            if not s.empty: df_metrics = pd.concat([df_metrics, s], ignore_index=True)

    df_visitors = pd.DataFrame()
    for f in LINKEDIN_RAW_DIR.glob("*manningshk_visitors*"):
        if f.suffix in ['.xlsx', '.xls']:
            s = smart_read_linkedin_excel(f, sheet_keyword='visitor')
            if not s.empty: df_visitors = pd.concat([df_visitors, s], ignore_index=True)

    if df_metrics.empty or df_visitors.empty: return pd.DataFrame()

    m_date_col = next((c for c in df_metrics.columns if 'date' in str(c).lower()), df_metrics.columns[0])
    v_date_col = next((c for c in df_visitors.columns if 'date' in str(c).lower()), df_visitors.columns[0])

    m_fmt, m_m, m_y, m_d = process_linkedin_dates(df_metrics[m_date_col], col_name=f"Metrics Base: {m_date_col}")
    df_metrics['_fmt_date'], df_metrics['_m'], df_metrics['_y'], df_metrics['_d'] = m_fmt, m_m, m_y, m_d
    v_fmt, v_m, v_y, v_d = process_linkedin_dates(df_visitors[v_date_col], col_name=f"Visitors Base: {v_date_col}")
    df_visitors['_fmt_date'], df_visitors['_m'], df_visitors['_y'], df_visitors['_d'] = v_fmt, v_m, v_y, v_d

    df_metrics = df_metrics[(df_metrics['_y'] == TARGET_YEAR) & (df_metrics['_m'] == TARGET_MONTH)].drop_duplicates(subset=['_fmt_date'])
    df_visitors = df_visitors[(df_visitors['_y'] == TARGET_YEAR) & (df_visitors['_m'] == TARGET_MONTH)].drop_duplicates(subset=['_fmt_date'])
    if df_metrics.empty or df_visitors.empty: return pd.DataFrame()

    df_merged = pd.merge(df_metrics, df_visitors, on='_fmt_date', how='outer', suffixes=('', '_v'))
    if '_m' not in df_merged.columns or df_merged['_m'].isnull().all(): df_merged['_m'] = df_merged['_m_v']
    if '_y' not in df_merged.columns or df_merged['_y'].isnull().all(): df_merged['_y'] = df_merged['_y_v']
    if '_d' not in df_merged.columns or df_merged['_d'].isnull().all(): df_merged['_d'] = df_merged['_d_v']

    # Chronological Sort Ascending
    df_merged = df_merged.sort_values(by=['_y', '_m', '_d'], ascending=True).copy()

    def get_val(df, keywords):
        c = next((col for col in df.columns if any(kw in str(col).lower() for kw in keywords)), None)
        return pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int) if c else pd.Series(0, index=df.index)

    df_out = pd.DataFrame()
    df_out['Month'] = df_merged['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
    df_out['Date'] = df_merged['_fmt_date']
    df_out['Desktop Views'] = get_val(df_merged, ['page views (desktop)'])
    df_out['Mobile Views'] = get_val(df_merged, ['page views (mobile)'])
    df_out['Total Page Views'] = get_val(df_merged, ['page views (total)'])
    df_out['Likes'] = get_val(df_merged, ['reactions (total)'])
    df_out['Comments'] = get_val(df_merged, ['comments (total)'])
    df_out['Shares'] = get_val(df_merged, ['reposts (total)', 'shares (total)'])
    df_out['Clicks'] = get_val(df_merged, ['clicks(total)', 'clicks (total)'])
    df_out['Impressions'] = get_val(df_merged, ['impressions (total)'])
    df_out['Likes (Organic)'] = get_val(df_merged, ['reactions (organic)'])
    df_out['Comments (Organic)'] = get_val(df_merged, ['comments (organic)'])
    df_out['Shares (Organic)'] = get_val(df_merged, ['reposts (organic)', 'shares (organic)'])
    df_out['Clicks (Organic)'] = get_val(df_merged, ['clicks(organic)', 'clicks (organic)'])
    df_out['Impressions (Organic)'] = get_val(df_merged, ['impressions (organic)'])
    df_out['Likes (Paid)'] = get_val(df_merged, ['reactions (sponsored)'])
    df_out['Comments (Paid)'] = get_val(df_merged, ['comments (sponsored)'])
    df_out['Shares (Paid)'] = get_val(df_merged, ['reposts (sponsored)', 'shares (sponsored)'])
    df_out['Clicks (Paid)'] = get_val(df_merged, ['clicks(sponsored)', 'clicks (sponsored)'])
    df_out['Impressions (Paid)'] = get_val(df_merged, ['impressions (sponsored)'])
    return df_out


def build_linkedin_follower_log():
    logging.info("Building LinkedIn Follower Log tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()
    log_date_str = pd.Timestamp.now().strftime('%m/%d/%Y')

    if TARGET_YEAR == 2026 and TARGET_MONTH < 7:
        f_log = LINKEDIN_RAW_DIR / "Follower log_202601_06.xlsx"
        if f_log.exists():
            try:
                df_log = smart_read_linkedin_excel(f_log)
                d_col = next((c for c in df_log.columns if 'date' in str(c).lower()), df_log.columns[0])
                fmt, mm, yy, dd = process_linkedin_dates(df_log[d_col], col_name=f"Hist Follower Log: {d_col}")
                df_log['_fmt_date'], df_log['_m'], df_log['_y'], df_log['_d'] = fmt, mm, yy, dd
                df_filtered = df_log[(df_log['_y'] == TARGET_YEAR) & (df_log['_m'] == TARGET_MONTH)].copy()
                if not df_filtered.empty:
                    df_filtered = df_filtered.sort_values(by=['_y', '_m', '_d'], ascending=True).copy()
                    df_out = pd.DataFrame()
                    df_out['Month'] = df_filtered['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
                    df_out['Date Polled'] = df_filtered['_fmt_date']
                    for col in ['Category', 'Demographic', 'Total Followers']:
                        match = next((c for c in df_filtered.columns if str(c).lower() == col.lower()), None)
                        df_out[col] = df_filtered[match] if match else ""
                    df_out['Log Date'] = log_date_str
                    return df_out
            except Exception as e: logging.error(f"  Error slicing historical file log: {e}")

    f_files = list(LINKEDIN_RAW_DIR.glob("*manningshk_followers*"))
    if not f_files: return pd.DataFrame()

    all_rows = []
    for f in f_files:
        try:
            xl = pd.ExcelFile(f)
            sheet = next((s for s in xl.sheet_names if 'followers' in s.lower()), None)
            if not sheet: continue
            df_nf = xl.parse(sheet, header=None)
            
            header_row_idx = 0
            for i, row in df_nf.iterrows():
                row_strs = [str(x).lower().strip() for x in row.dropna().values]
                if len(row_strs) > 1 and any(s in ['date', 'time'] for s in row_strs):
                    header_row_idx = i
                    break
            df_nf.columns = df_nf.iloc[header_row_idx]
            df_nf = df_nf.iloc[header_row_idx+1:].reset_index(drop=True)
            df_nf.columns = df_nf.columns.str.strip()
            
            d_col = next((c for c in df_nf.columns if 'date' in str(c).lower()), df_nf.columns[0])
            fmt, mm, yy, dd = process_linkedin_dates(df_nf[d_col], col_name=f"Daily Followers Log ({f.name}): {d_col}")
            df_nf['_fmt_date'], df_nf['_m'], df_nf['_y'], df_nf['_d'] = fmt, mm, yy, dd
            
            df_nf = df_nf[(df_nf['_y'] == TARGET_YEAR) & (df_nf['_m'] == TARGET_MONTH)].sort_values(by=['_y', '_m', '_d'], ascending=True)
            if df_nf.empty: continue

            org_c = next((c for c in df_nf.columns if 'organic' in str(c).lower()), None)
            paid_c = next((c for c in df_nf.columns if 'paid' in str(c).lower() or 'sponsored' in str(c).lower()), None)

            net_rolling = 0
            for _, r in df_nf.iterrows():
                m_str = MONTH_MAP[int(r['_m'])] if pd.notna(r['_m']) else ""
                org_v = pd.to_numeric(r[org_c], errors='coerce').fillna(0).astype(int) if org_c else 0
                paid_v = pd.to_numeric(r[paid_c], errors='coerce').fillna(0).astype(int) if paid_c else 0
                net_rolling += (org_v + paid_v)

                all_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Organic', 'Demographic': '', 'Total Followers': org_v, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})
                all_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Paid', 'Demographic': '', 'Total Followers': paid_v, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})
                all_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Net', 'Demographic': '', 'Total Followers': net_rolling, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})

            demo_categories = {'Location': ['location'], 'Job function': ['function'], 'Seniority': ['seniority'], 'Industry': ['industry'], 'Company size': ['size']}
            for cat_name, keywords in demo_categories.items():
                s_name = next((s for s in xl.sheet_names if any(kw in s.lower() for kw in keywords)), None)
                if s_name:
                    df_demo = xl.parse(s_name)
                    cnt_col = next((c for c in df_demo.columns if any(kw in c.lower() for kw in ['follower', 'count', 'total'])), df_demo.columns[1])
                    df_demo[cnt_col] = pd.to_numeric(df_demo[cnt_col], errors='coerce').fillna(0)
                    
                    last_row = df_nf.iloc[-1]
                    for _, d_row in df_demo.sort_values(by=cnt_col, ascending=False).head(5).iterrows():
                        all_rows.append({
                            'Month': MONTH_MAP[int(last_row['_m'])], 'Date Polled': last_row['_fmt_date'],
                            'Category': cat_name, 'Demographic': str(d_row[df_demo.columns[0]]), 'Total Followers': int(d_row[cnt_col]), 'Log Date': log_date_str,
                            '_y': last_row['_y'], '_m': last_row['_m'], '_d': last_row['_d']
                        })
        except Exception as e: logging.error(f"  Error calculating follower daily metrics: {e}")
        
    if all_rows:
        df_res = pd.DataFrame(all_rows)
        df_res = df_res.sort_values(by=['_y', '_m', '_d'], ascending=True).drop(columns=['_y', '_m', '_d']).reset_index(drop=True)
        return df_res
    return pd.DataFrame()

# ==============================================================================
# OUTPUT FRAME CACHE
# ==============================================================================
output_frames = {}

EXPECTED_TABS = [
    "FB Pivot (Category)",
    "FB Pivot (Post Type)",
    "FB Pivot (Pillar)",
    "FB Key Metrics",
    "FB Key Metric (CM)",
    "FB Wall Post Performance",
    "Category Performance - BAU",
    "Sub-Category Performance - PNP",
    "Pillar Performance - CRM",
    "Pillar Performance - Ecommerce",
    "Pillar Performance - GNC",
    "Pillar Performance - Others",
    "IG Story Performance",
    "IG Story Pivot",
    "IG Wall Post Performance",
    "IG Pivot (Pillar)",
    "IG Key Metrics",
    "IG Key Metric (CM)",
    "Master Comments Base",
    "Sentiment Summary (Category)",
    "Sentiment Summary (Type)",
    "LinkedIn Posts Perf",
    "LinkedIn Page Perf",
    "LinkedIn Follower Log"
]

def write_to_sheet(tab_name, dataframe):
    if dataframe is None:
        dataframe = pd.DataFrame()
    output_frames[tab_name] = dataframe.copy()
    if dataframe.empty:
        logging.info(f"  Tab '{tab_name}': empty (placeholder will be created)")
    else:
        logging.info(f"  Cached tab '{tab_name}': {len(dataframe)} rows")

# ==============================================================================
# MAIN PIPELINE
# ==============================================================================
def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info(f"=== 3_v5_period: Period Aggregation for {TARGET_PERIOD} ===")

    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_TRACKER_ID)
        logging.info(f"Connected to Social Tracker Google Sheet for {TARGET_PERIOD}.")
    except Exception as e:
        logging.error(f"Failed to connect: {e}")
        return

    # ------------------------------------------------------------------
    # 0. FETCH FOLLOWER DATA FROM API LOGS
    # ------------------------------------------------------------------
    fb_followers = fetch_followers_for_period("FB API Log")
    ig_followers = fetch_followers_for_period("IG API Log")

    # ------------------------------------------------------------------
    # 1. FACEBOOK DATA
    # ------------------------------------------------------------------
    logging.info("Processing FB data (All FB Posts_New Cat) ...")
    try:
        fb_ws = sh.worksheet("All FB Posts_New Cat")
        df_fb = pd.DataFrame(fb_ws.get_all_records())
        if not df_fb.empty:
            df_fb = normalize_columns(df_fb)
            df_fb_filtered = filter_to_target_period(df_fb, 'Publish time', dayfirst_guess=False)

            if len(df_fb_filtered) > 0:
                # --- FB Key Metrics (before dark post removal) ---
                try:
                    df_fb_key = generate_fb_key_metrics(df_fb_filtered.copy(), TARGET_PERIOD)
                    if not df_fb_key.empty:
                        df_fb_key = add_followers_to_key_metrics(df_fb_key, fb_followers)
                        write_to_sheet("FB Key Metrics", df_fb_key)
                except Exception as e:
                    logging.error(f"FB Key Metrics error: {e}")

                # --- FB Key Metric (CM) ---
                try:
                    df_fb_key_cm = generate_fb_key_metrics_cm(df_fb_filtered.copy(), TARGET_PERIOD)
                    if not df_fb_key_cm.empty:
                        df_fb_key_cm = add_followers_to_key_metrics(df_fb_key_cm, fb_followers)
                        write_to_sheet("FB Key Metric (CM)", df_fb_key_cm)
                except Exception as e:
                    logging.error(f"FB Key Metric (CM) error: {e}")

                # --- Remove dark posts for wall post tabs ---
                if 'Dark Post' in df_fb_filtered.columns:
                    df_fb_filtered = df_fb_filtered[df_fb_filtered['Dark Post'] != 'Y']
                if 'Month' in df_fb_filtered.columns:
                    df_fb_filtered = df_fb_filtered.drop(columns=['Month'])

                df_fb_filtered['Interactions_Numeric'] = clean_numeric_col(df_fb_filtered, 'Interactions')
                fb_pillar_col = get_actual_col_name(df_fb_filtered, ['Pillar'])
                fb_category_col = get_actual_col_name(df_fb_filtered, ['Category', 'category'])
                fb_post_type_col = get_actual_col_name(df_fb_filtered, ['Post type', 'post type', 'Post Type'])

                # Normalize Post Type naming (Photo→Photos, Video→Videos, etc.)
                if fb_post_type_col and fb_post_type_col in df_fb_filtered.columns:
                    df_fb_filtered[fb_post_type_col] = df_fb_filtered[fb_post_type_col].astype(str).str.strip().replace({
                        'Photo': 'Photos', 'Single Photo': 'Photos',
                        'Video': 'Videos', 'Link': 'Links',
                    })
                fb_subcat_col = get_actual_col_name(df_fb_filtered, ['Sub-Category', 'sub-category', 'Sub Category'])

                pivot_fb_category = df_fb_filtered.groupby(['_Period', fb_category_col]).agg(
                    No_of_post=(fb_category_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_category_col: 'Category',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                pivot_fb_post_type = df_fb_filtered.groupby(['_Period', fb_post_type_col]).agg(
                    No_of_post=(fb_post_type_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_post_type_col: 'Post type',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                pivot_fb_pillar = df_fb_filtered.groupby(['_Period', fb_pillar_col]).agg(
                    No_of_post=(fb_pillar_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_pillar_col: 'Pillar',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                write_to_sheet("FB Pivot (Category)", pivot_fb_category)
                write_to_sheet("FB Pivot (Post Type)", pivot_fb_post_type)
                write_to_sheet("FB Pivot (Pillar)", pivot_fb_pillar)

                df_fb_filtered = df_fb_filtered.drop(columns=['Interactions_Numeric', '_Period'])
                df_fb_filtered.columns = df_fb_filtered.columns.str.strip()

                fb_perf_cols = [
                    '_Year', '_Month', '_Day',
                    'Permalink', 'Description', 'Post type', 'Pillar', 'Category',
                    'Sub-Category', 'Campaign Name', 'Publish time', 'Reactions',
                    'Comments', 'Shares', 'Interactions', 'Reach', 'Views',
                    'Link Clicks', 'Photo Clicks', '3-second video views',
                    'Video Length', 'Organic reach', 'Paid reach', 'Dark Post',
                    'Organic', 'Paid', 'Month No.', 'Day', 'Hour', 'Average Watch Time'
                ]
                fb_numeric_cols = [
                    'Reactions', 'Comments', 'Shares', 'Interactions', 'Reach', 'Views',
                    'Link Clicks', 'Photo Clicks', '3-second video views', 'Organic reach',
                    'Paid reach', 'Month No.', 'Day', 'Hour', 'Video Length', 'Average Watch Time'
                ]

                def prep_fb_perf(df_subset):
                    if df_subset is None or len(df_subset) == 0:
                        return pd.DataFrame(columns=fb_perf_cols)
                    result = df_subset.copy()
                    for col in fb_perf_cols:
                        if col not in result.columns:
                            result[col] = ""
                    result = result[fb_perf_cols]
                    for col in fb_numeric_cols:
                        if col in result.columns:
                            result[col] = pd.to_numeric(result[col].astype(str).str.replace(',', ''), errors='coerce')
                    return result

                write_to_sheet("FB Wall Post Performance", prep_fb_perf(df_fb_filtered))

                perf_breakdown_cols = ['_Year', '_Month', '_Day', 'Permalink', 'Description', 'Post type', 'Pillar', 'Category', 'Sub-Category',
                                       'Campaign Name', 'Publish time', 'Reactions', 'Comments', 'Shares',
                                       'Interactions', 'Reach', 'Views', 'Link Clicks', 'Photo Clicks']

                def prep_breakdown(df_subset):
                    df_out = prep_fb_perf(df_subset)
                    cols = [c for c in perf_breakdown_cols if c in df_out.columns]
                    return df_out[cols] if cols else df_out

                cat_clean = df_fb_filtered['Category'].astype(str).str.lower().str.strip() if 'Category' in df_fb_filtered.columns else pd.Series(index=df_fb_filtered.index).fillna("")
                subcat_clean = df_fb_filtered[fb_subcat_col].astype(str).str.lower().str.strip() if fb_subcat_col in df_fb_filtered.columns else pd.Series(index=df_fb_filtered.index).fillna("")
                write_to_sheet("Category Performance - BAU", prep_breakdown(df_fb_filtered[cat_clean == 'bau promotion']))
                write_to_sheet("Sub-Category Performance - PNP", prep_breakdown(df_fb_filtered[subcat_clean == 'friday pnp']))
                write_to_sheet("Pillar Performance - CRM", prep_breakdown(df_fb_filtered[cat_clean.str.contains('yuu', na=False) | (cat_clean == 'weekly bau')]))
                write_to_sheet("Pillar Performance - Ecommerce", prep_breakdown(df_fb_filtered[cat_clean.str.contains('ecom', na=False)]))
                write_to_sheet("Pillar Performance - GNC", prep_breakdown(df_fb_filtered[cat_clean == 'gnc']))
                write_to_sheet("Pillar Performance - Others", prep_breakdown(df_fb_filtered[cat_clean == 'others']))
    except Exception as e:
        logging.error(f"FB processing error: {e}")

    # ------------------------------------------------------------------
    # 2. IG STORY DATA
    # ------------------------------------------------------------------
    logging.info("Processing IG Story data (All_IGS) ...")
    df_igs = None
    df_igs_filtered = pd.DataFrame()
    try:
        igs_ws = sh.worksheet("All_IGS")
        df_igs = pd.DataFrame(igs_ws.get_all_records())
        if not df_igs.empty:
            df_igs = normalize_columns(df_igs)
            df_igs_filtered = filter_to_target_period(df_igs, 'Publish time', dayfirst_guess=True)
            if len(df_igs_filtered) > 0:
                igs_pillar_col = get_actual_col_name(df_igs_filtered, ['Pillar'])
                link_click_col = get_actual_col_name(df_igs_filtered, ['Link clicks', 'link clicks'])
                df_igs_filtered['Link_Clicks_Numeric'] = clean_numeric_col(df_igs_filtered, link_click_col)

                pivot_igs_pillar = df_igs_filtered.groupby(['_Period', igs_pillar_col]).agg(
                    No_of_post=(igs_pillar_col, 'count'),
                    Total_Link_Clicks=('Link_Clicks_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', igs_pillar_col: 'Pillar',
                    'No_of_post': 'No. of Posts', 'Total_Link_Clicks': 'SUM of Link clicks'})
                df_igs_filtered = df_igs_filtered.drop(columns=['Link_Clicks_Numeric', '_Period'])

                igs_cols_to_keep = ['_Year', '_Month', '_Day', 'Publish time', 'Description', 'Pillar', 'Category', 'Total Reach', 'Shares', 'Link clicks']
                igs_cols_actual = [get_actual_col_name(df_igs_filtered, [c, c.lower(), c.capitalize()]) for c in igs_cols_to_keep]
                df_igs_filtered = df_igs_filtered[[c for c in igs_cols_actual if c in df_igs_filtered.columns]]
                for col in ['Total Reach', 'Shares', 'Link clicks']:
                    actual = get_actual_col_name(df_igs_filtered, [col, col.lower()])
                    if actual in df_igs_filtered.columns:
                        df_igs_filtered[actual] = pd.to_numeric(
                            df_igs_filtered[actual].astype(str).str.replace(',', ''), errors='coerce')

                write_to_sheet("IG Story Performance", df_igs_filtered)
                write_to_sheet("IG Story Pivot", pivot_igs_pillar)
    except Exception as e:
        logging.error(f"IG Story error: {e}")

    # ------------------------------------------------------------------
    # 3. IG WALL POST DATA
    # ------------------------------------------------------------------
    logging.info("Processing IG Wall Post data (All_IG_Posts) ...")
    df_igp = None
    df_igp_filtered = pd.DataFrame()
    try:
        igp_ws = sh.worksheet("All_IG_Posts")
        df_igp = pd.DataFrame(igp_ws.get_all_records())
        if not df_igp.empty:
            df_igp = normalize_columns(df_igp)
            df_igp_filtered = filter_to_target_period(df_igp, 'Publish time', dayfirst_guess=True)
            if len(df_igp_filtered) > 0:
                pub_time_col = get_actual_col_name(df_igp_filtered, ['Publish time', 'publish time'])
                views_col = get_actual_col_name(df_igp_filtered, ['Views', 'views'])
                df_igp_filtered.rename(columns={pub_time_col: 'Post Date', views_col: 'Impressions'}, inplace=True)

                final_igp_cols = [
                    '_Year', '_Month', '_Day',
                    'Post Date', 'Type', 'Post Message', 'Pillar', 'Category',
                    'Sub-Category', 'Campaign Name', 'Total Post Reach',
                    'Impressions', 'Likes', 'Comments', 'Shares', 'Saves', 'Total Interactions'
                ]
                for col in final_igp_cols:
                    if col not in df_igp_filtered.columns:
                        df_igp_filtered[col] = ""

                df_igp_filtered['Total_Interactions_Numeric'] = clean_numeric_col(df_igp_filtered, 'Total Interactions')

                pivot_igp_pillar = df_igp_filtered.groupby(['_Period', 'Pillar']).agg(
                    No_of_post=('Pillar', 'count'),
                    Sum_Interactions=('Total_Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period',
                    'No_of_post': 'No. of Posts', 'Sum_Interactions': 'Total Interactions'})

                df_igp_filtered = df_igp_filtered.drop(columns=['Total_Interactions_Numeric', '_Period'])
                df_igp_filtered = df_igp_filtered[final_igp_cols]
                for col in ['Total Post Reach', 'Impressions', 'Likes', 'Comments', 'Shares', 'Saves', 'Total Interactions']:
                    if col in df_igp_filtered.columns:
                        df_igp_filtered[col] = pd.to_numeric(
                            df_igp_filtered[col].astype(str).str.replace(',', ''), errors='coerce')

                write_to_sheet("IG Wall Post Performance", df_igp_filtered)
                write_to_sheet("IG Pivot (Pillar)", pivot_igp_pillar)
    except Exception as e:
        logging.error(f"IG Wall Post error: {e}")

    # ------------------------------------------------------------------
    # 4. IG KEY METRICS
    # ------------------------------------------------------------------
    try:
        ig_post_for_metrics = df_igp_filtered if not df_igp_filtered.empty else None
        ig_story_for_metrics = df_igs_filtered if not df_igs_filtered.empty else None
        df_ig_key = generate_ig_key_metrics(ig_post_for_metrics, ig_story_for_metrics)
        if not df_ig_key.empty:
            df_ig_key = add_followers_to_key_metrics(df_ig_key, ig_followers)
            write_to_sheet("IG Key Metrics", df_ig_key)
    except Exception as e:
        logging.error(f"IG Key Metrics error: {e}")

    # ------------------------------------------------------------------
    # 4b. IG KEY METRIC (CM)
    # ------------------------------------------------------------------
    try:
        ig_post_for_metrics_cm = df_igp_filtered if not df_igp_filtered.empty else None
        ig_story_for_metrics_cm = df_igs_filtered if not df_igs_filtered.empty else None
        df_ig_key_cm = generate_ig_key_metrics_cm(ig_post_for_metrics_cm, ig_story_for_metrics_cm)
        if not df_ig_key_cm.empty:
            df_ig_key_cm = add_followers_to_key_metrics(df_ig_key_cm, ig_followers)
            write_to_sheet("IG Key Metric (CM)", df_ig_key_cm)
    except Exception as e:
        logging.error(f"IG Key Metric (CM) error: {e}")

    # ------------------------------------------------------------------
    # 5. COMMENTS (from period-specific cleaned Excel file)
    # ------------------------------------------------------------------
    if os.path.exists(COMMENTS_XLSX):
        logging.info(f"Loading cleaned comments: {COMMENTS_XLSX.name} ...")
        df_comments = pd.read_excel(COMMENTS_XLSX)
        if 'Post Date' in df_comments.columns:
            parsed_comment_dates = smart_parse_dates(df_comments['Post Date'].astype(str).str.strip())
            df_comments['Post Date'] = parsed_comment_dates.dt.strftime('%Y-%m-%d %H:%M')
        if 'Comment Date' in df_comments.columns:
            month_hints = df_comments['Month Note'] if 'Month Note' in df_comments.columns else None
            parsed_cd = smart_parse_dates(df_comments['Comment Date'].astype(str).str.strip(), month_hints=month_hints)
            df_comments['_Period'] = parsed_cd.dt.strftime('%Y-%m')
            df_comments['_Year'] = parsed_cd.dt.year.fillna(0).astype(int)
            df_comments['_Month'] = parsed_cd.dt.month.fillna(0).astype(int)
            df_comments['_Day'] = parsed_cd.dt.day.fillna(0).astype(int)
        elif 'Post Date' in df_comments.columns:
            df_comments['_Period'] = TARGET_PERIOD
            df_comments['_Year'] = parsed_comment_dates.dt.year.fillna(0).astype(int)
            df_comments['_Month'] = parsed_comment_dates.dt.month.fillna(0).astype(int)
            df_comments['_Day'] = parsed_comment_dates.dt.day.fillna(0).astype(int)
        else:
            df_comments['_Period'] = TARGET_PERIOD
            df_comments['_Year'] = TARGET_YEAR
            df_comments['_Month'] = TARGET_MONTH
            df_comments['_Day'] = 0
        write_to_sheet("Master Comments Base", df_comments)

        if not df_comments.empty and 'Sentiment' in df_comments.columns:
            df_comments['Sentiment'] = df_comments['Sentiment'].astype(str).str.strip().str.title().replace('N/a', 'N/A')
            df_comments = df_comments[df_comments['Sentiment'].isin(['Positive', 'Neutral', 'Negative', 'N/A'])]

            if 'Category' in df_comments.columns and '_Period' in df_comments.columns:
                sent_cat_pivot = df_comments.groupby(['_Period', 'Category', 'Sentiment']).size().unstack(fill_value=0).reset_index()
                for s in ['Positive', 'Neutral', 'Negative', 'N/A']:
                    if s not in sent_cat_pivot.columns:
                        sent_cat_pivot[s] = 0
                sent_cat_pivot['Total'] = sent_cat_pivot[['Positive', 'Neutral', 'Negative', 'N/A']].sum(axis=1)
                sent_cat_pivot = sent_cat_pivot.sort_values(['_Period', 'Total'], ascending=[True, False])
                sent_cat_pivot = sent_cat_pivot.rename(columns={'_Period': 'Period'})
                period_col = sent_cat_pivot.pop('Period')
                sent_cat_pivot.insert(0, 'Period', period_col)
                write_to_sheet("Sentiment Summary (Category)", sent_cat_pivot)

            type_col = 'Type' if 'Type' in df_comments.columns else ('type' if 'type' in df_comments.columns else None)
            if type_col and '_Period' in df_comments.columns:
                sent_type_pivot = df_comments.groupby(['_Period', type_col, 'Sentiment']).size().unstack(fill_value=0).reset_index()
                for s in ['Positive', 'Neutral', 'Negative', 'N/A']:
                    if s not in sent_type_pivot.columns:
                        sent_type_pivot[s] = 0
                sent_type_pivot['Total'] = sent_type_pivot[['Positive', 'Neutral', 'Negative', 'N/A']].sum(axis=1)
                sent_type_pivot = sent_type_pivot.sort_values(['_Period', 'Total'], ascending=[True, False])
                sent_type_pivot = sent_type_pivot.rename(columns={'_Period': 'Period'})
                period_col = sent_type_pivot.pop('Period')
                sent_type_pivot.insert(0, 'Period', period_col)
                write_to_sheet("Sentiment Summary (Type)", sent_type_pivot)
    else:
        logging.warning(f"Comments Excel file not found: {COMMENTS_XLSX}")

    # ------------------------------------------------------------------
    # 5b. ADD API-SOURCED TABS + REACH FUNNEL
    # ------------------------------------------------------------------
    api_tabs = build_api_tabs_period()
    for tab_name, df_api in api_tabs.items():
        write_to_sheet(tab_name, df_api)

    df_reach_funnel = build_reach_funnel_tab_period()
    if df_reach_funnel is not None and not df_reach_funnel.empty:
        write_to_sheet("FB Reach Funnel", df_reach_funnel)

    # ------------------------------------------------------------------
    # 5c. LINKEDIN DATA PIPELINE
    # ------------------------------------------------------------------
    logging.info("Processing LinkedIn data tabs ...")
    try:
        df_li_posts = build_linkedin_posts_perf()
        write_to_sheet("LinkedIn Posts Perf", df_li_posts)

        df_li_page = build_linkedin_page_perf()
        write_to_sheet("LinkedIn Page Perf", df_li_page)

        df_li_followers = build_linkedin_follower_log()
        write_to_sheet("LinkedIn Follower Log", df_li_followers)
    except Exception as e:
        logging.error(f"LinkedIn data aggregation error: {e}")

    # ------------------------------------------------------------------
    # 5d. BACKFILL MISSING TABS
    # ------------------------------------------------------------------
    for tab in EXPECTED_TABS:
        if tab not in output_frames or output_frames[tab].empty:
            placeholder = pd.DataFrame({
                'Period': [TARGET_PERIOD],
                'Note': [f'No data available for {TARGET_PERIOD}'],
            })
            output_frames[tab] = placeholder
            logging.info(f"  Backfilled empty tab: '{tab}'")

    # ------------------------------------------------------------------
    # 6. EXPORT TO EXCEL
    # ------------------------------------------------------------------
    logging.info(f"Exporting to Excel: {FEED_OUTPUT}")
    if output_frames:
        API_TAB_ORDER = ["Unique Page View", "FB Followers", "IG Followers", "FB Reach Funnel"]
        ordered_sheets = [t for t in EXPECTED_TABS if t in output_frames]
        ordered_sheets += [t for t in API_TAB_ORDER if t in output_frames]
        extras = [t for t in output_frames.keys() if t not in EXPECTED_TABS and t not in API_TAB_ORDER]
        ordered_sheets += extras

        logging.info(f"Exporting {len(ordered_sheets)} sheets to Excel: {FEED_OUTPUT}")
        with pd.ExcelWriter(str(FEED_OUTPUT), engine='openpyxl') as writer:
            for sheet_name in ordered_sheets:
                df_out = output_frames[sheet_name]
                # Avoid standard normalization on LinkedIn logs to preserve strict m/d/yyyy format
                if sheet_name not in ["LinkedIn Posts Perf", "LinkedIn Page Perf", "LinkedIn Follower Log"]:
                    df_out = normalize_date_columns(df_out)
                clean_df = deep_clean_df(df_out.fillna(""))
                safe_name = sheet_name[:31]
                clean_df.to_excel(writer, sheet_name=safe_name, index=False)

            from openpyxl.styles import Alignment

            INT_KW = ['reach', 'likes', 'comment', 'share', 'save', 'interaction',
                      'impression', 'click', 'follower', 'growth', 'no. of',
                      'posted', 'paid', 'wall post', 'view', 'reactions',
                      'sum of link', 'count', 'dark post', 'organic', 'video',
                      'feeds', 'reposts', 'follows', 'desktop', 'mobile', 'total']
            PCT_KW = ['rate', '%']
            DP1_KW = ['average']
            PCT0_KW = ['organic %', 'paid %']

            def col_fmt(header):
                h = str(header).lower()
                if any(kw in h for kw in PCT0_KW):
                    return '0%'
                if any(kw in h for kw in PCT_KW):
                    return '0.00%'
                if any(kw in h for kw in DP1_KW):
                    return '#,##0.0'
                if any(kw in h for kw in INT_KW):
                    return '#,##0'
                return None

            def row_fmt(metric_label):
                h = str(metric_label).lower()
                if any(kw in h for kw in PCT0_KW):
                    return '0%'
                if any(kw in h for kw in PCT_KW):
                    return '0.00%'
                if any(kw in h for kw in DP1_KW):
                    return '#,##0.0'
                if any(kw in h for kw in INT_KW):
                    return '#,##0'
                return None

            ROW_FMT_TABS = ['FB Key Metrics', 'FB Key Metric (CM)', 'IG Key Metrics', 'IG Key Metric (CM)']

            for ws in writer.book.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        cur = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=cur.horizontal,
                            vertical=cur.vertical,
                            wrap_text=False,
                        )

                if ws.title in ROW_FMT_TABS:
                    for row in ws.iter_rows(min_row=2, min_col=1):
                        label = row[0].value
                        fmt = row_fmt(label)
                        if fmt:
                            for cell in row[1:]:
                                cell.number_format = fmt
                else:
                    headers = {}
                    for cell in ws[1]:
                        headers[cell.column] = cell.value
                    for col_idx, header in headers.items():
                        fmt = col_fmt(header)
                        if fmt:
                            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    cell.number_format = fmt

        logging.info(f"DONE! Excel exported: {FEED_OUTPUT}")
        logging.info(f"Total tabs: {len(output_frames)}")
        for name in output_frames.keys():
            logging.info(f"  - {name}")
    else:
        logging.warning("No data to export.")


if __name__ == "__main__":
    main()