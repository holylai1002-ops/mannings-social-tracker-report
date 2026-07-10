"""
3_new_v3.py - Master Aggregation for ALL-Period Dashboard (Feed output)

Based on 3_v5_period.py but processes ALL periods (no single-month filter).
  - FB Key Metrics + FB Key Metric (CM) tabs (periods as columns)
    incl. Total Followers + Net Followers Growth from FB API Log
  - IG Key Metrics + IG Key Metric (CM) tabs (periods as columns)
    incl. Total Followers + Net Followers Growth from IG API Log
  - Keeps dark posts in Key Metrics tabs
  - Reads comments from xlsx (not csv)
  - EXPECTED_TABS backfill for empty tabs
  - API-sourced tabs (Unique Page View, FB Followers, IG Followers)
  - FB Reach Funnel tab
  - Added LinkedIn tabs (Posts Perf, Page Perf, Follower Log) processing ALL periods
  - Chronological date sorting (Ascending) for LinkedIn tabs without target filters
  - Fixed LinkedIn Follower Log demographics date assignment and rolling Net accumulation
  - Fixed 'int' object has no attribute 'fillna' scalar processing bug
  - Resolved TARGET_YEAR / TARGET_MONTH NameError by stripping period-specific logic
  - deep_clean_df for Excel-safe output

Output: Mannings_FB_IG_Dashboard_Feed.xlsx (final dashboard feed)
"""
import os
import re
import json
import datetime
import logging
from pathlib import Path
import pandas as pd

from config import BASE_DIR, GSHEET_TRACKER_ID, VALID_YEARS, FINAL_EXCEL_OUTPUT

FEED_OUTPUT = FINAL_EXCEL_OUTPUT
from utils import (
    get_gspread_client, clean_excel_characters, clean_numeric_col,
    get_actual_col_name, normalize_columns, smart_parse_dates
)

COMMENTS_XLSX = BASE_DIR / "FB Comments - ALL.xlsx"
IG_API_FILE = BASE_DIR / "ig_api_data.json"

GSHEET_KEY_METRICS_ID = "1f9HLS0HXs2B-a_fxvvcUKuRKz0iMvmEDVivoLFl2vr8"

# LinkedIn Data Configurations
LINKEDIN_RAW_DIR = Path(r"C:\Users\holylai\Documents\n8n\Mannings\LinkedIn raw")
MONTH_MAP = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ==============================================================================
# EXCEL-SAFE CLEANING
# ==============================================================================
_OPENPYXL_ILLEGAL_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff\ufffe\uffff]'
)

def strip_illegal_for_excel(val):
    if val is None:
        return ""
    if isinstance(val, (int, float, bool)):
        return val
    s = str(val)
    s = _OPENPYXL_ILLEGAL_RE.sub('', s)
    s = re.sub(r'[\x7f-\x9f\u200b-\u200f\u2028-\u202f\ufeff]', '', s)
    return s

def deep_clean_df(df):
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].apply(strip_illegal_for_excel)
    return df


# ==============================================================================
# STRATEGIC LINKEDIN HYBRID DATE PARSER
# ==============================================================================
def process_linkedin_dates(series, col_name="Date"):
    """
    Intelligently checks variable types (native datetime64, dash-strings, or slash-strings).
    Converts items to integers to drop leading zeros, producing clean m/d/yyyy strings.
    Returns day element keys to support comprehensive ascending sorting logic.
    """
    results = []
    months = []
    years = []
    days = []
    
    for idx, val in series.items():
        if pd.isna(val) or str(val).strip().lower() in ['nan', 'none', '']:
            results.append("")
            months.append(None)
            years.append(None)
            days.append(None)
            continue
            
        # Case 1: Pandas auto-parsed the column as a datetime64/Timestamp object
        if isinstance(val, (pd.Timestamp, datetime.date)) or hasattr(val, 'month'):
            m, d, y = val.month, val.day, val.year
            results.append(f"{m}/{d}/{y}")
            months.append(m)
            years.append(y)
            days.append(d)
            continue
            
        # Case 2: Raw String Fallback (Regex-based grouping for dashes/slashes)
        val_str = str(val).strip()
        match = re.search(r'(\d+)[/-](\d+)[/-](\d+)', val_str)
        if match:
            try:
                p1, p2, p3 = match.groups()
                if len(p1) == 4:  # ISO Format string: YYYY-MM-DD
                    y, m, d = int(p1), int(p2), int(p3)
                else:  # Custom Format string: MM/DD/YYYY or M/D/YYYY
                    m, d, y = int(p1), int(p2), int(p3)
                    
                results.append(f"{m}/{d}/{y}")
                months.append(m)
                years.append(y)
                days.append(d)
            except:
                results.append("")
                months.append(None)
                years.append(None)
                days.append(None)
        else:
            results.append("")
            months.append(None)
            years.append(None)
            days.append(None)
        
    return (
        pd.Series(results, index=series.index),
        pd.Series(months, index=series.index),
        pd.Series(years, index=series.index),
        pd.Series(days, index=series.index)
    )


# ==============================================================================
# LINKEDIN SMART SHEET INPUT MANAGEMENT ENGINE
# ==============================================================================
def smart_read_linkedin_excel(file_path, sheet_keyword=None):
    """
    Locates the true header block by checking for rows with multiple parsed columns,
    efficiently skipping any descriptive summary rows text lines exported from LinkedIn.
    """
    try:
        if isinstance(file_path, pd.ExcelFile):
            xl = file_path
        else:
            xl = pd.ExcelFile(file_path)
            
        if sheet_keyword:
            sheet_name = next((s for s in xl.sheet_names if sheet_keyword.lower() in s.lower()), xl.sheet_names[0])
        else:
            sheet_name = xl.sheet_names[0]
            
        df = xl.parse(sheet_name, header=None)
        
        header_row_idx = 0
        for i, row in df.iterrows():
            row_strs = [str(x).lower().strip() for x in row.dropna().values]
            if len(row_strs) > 3 and any(s in ['date', 'created date', 'post title', 'impressions', 'total page views', 'views', 'followers', 'category'] for s in row_strs):
                header_row_idx = i
                break
                
        df.columns = df.iloc[header_row_idx]
        df = df.iloc[header_row_idx+1:].reset_index(drop=True)
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        logging.error(f"Error smart-reading sheet {sheet_keyword}: {e}")
        return pd.DataFrame()


# ==============================================================================
# DATE NORMALIZATION
# ==============================================================================
_DATE_COL_NAMES = {'Publish time', 'Post Date', 'Comment Date', 'Date'}

def normalize_date_columns(df):
    """Normalize all date-like columns to DD/MM/YYYY HH:MM format."""
    if df is None or df.empty:
        return df
    df = df.copy()
    for col in df.columns:
        col_lower = str(col).lower().strip()
        is_date_col = (col in _DATE_COL_NAMES or
                       'date' in col_lower or
                       'publish time' in col_lower or
                       'post date' in col_lower)
        if not is_date_col:
            continue
        vals = df[col].astype(str).str.strip()
        non_empty = vals[(vals != '') & (vals != 'nan') & (vals != 'None')]
        if len(non_empty) == 0:
            continue
        parsed = smart_parse_dates(vals)
        has_time = non_empty.str.contains(':').any()
        if has_time:
            df[col] = parsed.dt.strftime('%d/%m/%Y %H:%M')
        else:
            df[col] = parsed.dt.strftime('%d/%m/%Y')
    return df


# ==============================================================================
# PERIOD ASSIGNMENT (multi-period)
# ==============================================================================
def assign_periods(df, date_col='Publish time'):
    if df is None or df.empty:
        return df
    df = df.copy()

    if date_col not in df.columns:
        return df

    month_hints = df['Month'] if 'Month' in df.columns else None
    parsed = smart_parse_dates(df[date_col].astype(str).str.strip(), month_hints=month_hints)
    years = parsed.dt.year

    month_no_col = None
    for candidate in ['Month No.', 'Month No', 'month no.']:
        if candidate in df.columns:
            month_no_col = candidate
            break

    if month_no_col:
        months = pd.to_numeric(df[month_no_col], errors='coerce')
        logging.info(f"  assign_periods: using '{month_no_col}' for month ({len(df)} rows)")
    else:
        months = parsed.dt.month
        logging.info(f"  assign_periods: using parsed dates for month ({len(df)} rows)")

    mask = years.isin(VALID_YEARS) & months.notna() & (months >= 1) & (months <= 12)
    df = df[mask].copy()
    df['_Period'] = (years[mask].astype(int).astype(str) + '-' +
                     months[mask].astype(int).astype(str).str.zfill(2)).values
    df['_Year'] = years[mask].astype(int).values
    df['_Month'] = months[mask].astype(int).values
    df['_Day'] = parsed[mask].dt.day.fillna(0).astype(int).values
    return df


# ==============================================================================
# FB KEY METRICS (multi-period)
# ==============================================================================
def generate_fb_key_metrics_all(df_fb):
    logging.info("Generating FB Key Metrics (all periods) ...")
    if df_fb is None or df_fb.empty or '_Period' not in df_fb.columns:
        logging.warning("  No FB data for Key Metrics")
        return pd.DataFrame()

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    def safe_count_val(df, col, val):
        if col in df.columns:
            return (df[col].astype(str).str.strip() == val).sum()
        return 0

    metrics_list = []
    metric_order = [
        'Feeds Posted', 'Total Dark Posts', 'Total Organic', 'Organic %',
        'Total Paid', 'Paid %', 'No. of Video Post', 'Total Interactions',
        'Total Reactions', 'Total Comments', 'Total Shares', 'Total Video Views',
        'Average Organic Reach', 'Average Paid Reach', 'Average Interaction',
        'Average Reactions', 'Average Comments', 'Average Shares', 'Average Video Views',
    ]

    for ym in sorted(df_fb['_Period'].unique()):
        ym_data = df_fb[df_fb['_Period'] == ym]

        feeds_posted = len(ym_data)
        total_dark_posts = safe_count_val(ym_data, 'Dark Post', 'Y')
        total_organic = safe_count_val(ym_data, 'Organic', 'Organic')
        total_paid = safe_count_val(ym_data, 'Paid', 'Paid')
        organic_pct = total_organic / feeds_posted if feeds_posted > 0 else 0
        paid_pct = total_paid / feeds_posted if feeds_posted > 0 else 0

        video_views_col = '3-second video views'
        no_video_posts = 0
        if video_views_col in ym_data.columns:
            col_raw = ym_data[video_views_col].astype(str).str.strip()
            has_content = col_raw.str.len() > 0
            vv = pd.to_numeric(col_raw, errors='coerce')
            no_video_posts = int((has_content & vv.notna()).sum())

        total_interactions = safe_sum(ym_data, 'Interactions')
        total_reactions = safe_sum(ym_data, 'Reactions')
        total_comments = safe_sum(ym_data, 'Comments')
        total_shares = safe_sum(ym_data, 'Shares')
        total_video_views = safe_sum(ym_data, '3-second video views')
        total_organic_reach = safe_sum(ym_data, 'Organic reach')
        total_paid_reach = safe_sum(ym_data, 'Paid reach')

        d = no_video_posts
        row_data = {
            'Feeds Posted': feeds_posted,
            'Total Dark Posts': total_dark_posts,
            'Total Organic': total_organic,
            'Organic %': organic_pct,
            'Total Paid': total_paid,
            'Paid %': paid_pct,
            'No. of Video Post': no_video_posts,
            'Total Interactions': int(total_interactions),
            'Total Reactions': int(total_reactions),
            'Total Comments': int(total_comments),
            'Total Shares': int(total_shares),
            'Total Video Views': int(total_video_views),
            'Average Organic Reach': round(total_organic_reach / d, 1) if d > 0 else 0,
            'Average Paid Reach': round(total_paid_reach / d, 1) if d > 0 else 0,
            'Average Interaction': round(total_interactions / d, 1) if d > 0 else 0,
            'Average Reactions': round(total_reactions / d, 1) if d > 0 else 0,
            'Average Comments': round(total_comments / d, 1) if d > 0 else 0,
            'Average Shares': round(total_shares / d, 1) if d > 0 else 0,
            'Average Video Views': round(total_video_views / d, 1) if d > 0 else 0,
        }

        for metric in metric_order:
            metrics_list.append({'Metric': metric, ym: row_data[metric]})

    if not metrics_list:
        return pd.DataFrame()

    df_metrics = pd.DataFrame(metrics_list)
    df_metrics = df_metrics.groupby('Metric', sort=False).first().reset_index()
    return df_metrics


# ==============================================================================
# FB KEY METRIC (CM) (multi-period)
# ==============================================================================
def generate_fb_key_metrics_cm_all(df_fb):
    logging.info("Generating FB Key Metric (CM) (all periods) ...")
    if df_fb is None or df_fb.empty or '_Period' not in df_fb.columns:
        return pd.DataFrame()

    if 'Pillar' in df_fb.columns:
        mask = ~df_fb['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_fb = df_fb[mask].copy()

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    def safe_count_val(df, col, val):
        if col in df.columns:
            return (df[col].astype(str).str.strip() == val).sum()
        return 0

    metrics_list = []
    metric_order = [
        'Feeds Posted', 'No. of Wall Post', 'No. of Dark Post',
        'No. of Organic Post', 'Organic Post %', 'No. of Paid Post', 'Paid Post %',
        'Total Interaction', 'Total Reactions', 'Total Comments', 'Total Shares',
        'Average Organic Reach', 'Average Paid Reach', 'Average Interaction',
        'Average Reactions', 'Average Comments', 'Average Shares',
    ]

    for ym in sorted(df_fb['_Period'].unique()):
        ym_data = df_fb[df_fb['_Period'] == ym]

        feeds_posted = len(ym_data)
        no_dark_post = safe_count_val(ym_data, 'Dark Post', 'Y')
        no_wall_post = feeds_posted - no_dark_post
        no_organic = safe_count_val(ym_data, 'Organic', 'Organic')
        no_paid = safe_count_val(ym_data, 'Paid', 'Paid')
        organic_pct = no_organic / feeds_posted if feeds_posted > 0 else 0
        paid_pct = no_paid / feeds_posted if feeds_posted > 0 else 0

        total_interactions = safe_sum(ym_data, 'Interactions')
        total_reactions = safe_sum(ym_data, 'Reactions')
        total_comments = safe_sum(ym_data, 'Comments')
        total_shares = safe_sum(ym_data, 'Shares')
        total_organic_reach = safe_sum(ym_data, 'Organic reach')
        total_paid_reach = safe_sum(ym_data, 'Paid reach')

        d = no_wall_post
        row_data = {
            'Feeds Posted': feeds_posted,
            'No. of Wall Post': no_wall_post,
            'No. of Dark Post': no_dark_post,
            'No. of Organic Post': no_organic,
            'Organic Post %': round(organic_pct, 4),
            'No. of Paid Post': no_paid,
            'Paid Post %': round(paid_pct, 4),
            'Total Interaction': int(total_interactions),
            'Total Reactions': int(total_reactions),
            'Total Comments': int(total_comments),
            'Total Shares': int(total_shares),
            'Average Organic Reach': round(total_organic_reach / d, 1) if d > 0 else 0,
            'Average Paid Reach': round(total_paid_reach / d, 1) if d > 0 else 0,
            'Average Interaction': round(total_interactions / d, 1) if d > 0 else 0,
            'Average Reactions': round(total_reactions / d, 1) if d > 0 else 0,
            'Average Comments': round(total_comments / d, 1) if d > 0 else 0,
            'Average Shares': round(total_shares / d, 1) if d > 0 else 0,
        }

        for metric in metric_order:
            metrics_list.append({'Metric': metric, ym: row_data[metric]})

    if not metrics_list:
        return pd.DataFrame()

    df_metrics = pd.DataFrame(metrics_list)
    df_metrics = df_metrics.groupby('Metric', sort=False).first().reset_index()
    return df_metrics


# ==============================================================================
# IG KEY METRICS (multi-period, updated order)
# ==============================================================================
def generate_ig_key_metrics_all(df_ig_post, df_ig_story):
    logging.info("Generating IG Key Metrics (all periods) ...")

    if df_ig_post is None or df_ig_post.empty or '_Period' not in df_ig_post.columns:
        logging.warning("  No IG Post data for Key Metrics")
        return pd.DataFrame()

    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').sum()
        return 0

    metrics_list = []
    metric_order = [
        'Feeds Posted', 'Total Post Likes', 'Total Post Comments', 'Total Post Shares',
        'Total Post Saves', 'Total Post Interaction', 'Total Post Reach',
        'Average Post Likes', 'Average Post Comments', 'Average Post Shares',
        'Average Post Saves', 'Average Post Interaction', 'Average Post Reach',
        'Stories Posted', 'Total Stories Link Clicks', 'Total Stories Shares',
        'Total Stories Reach', 'Average Stories Link Clicks', 'Average Stories Shares',
        'Average Stories Reach',
    ]

    for ym in sorted(df_ig_post['_Period'].unique()):
        ym_data = df_ig_post[df_ig_post['_Period'] == ym]
        total_posts = len(ym_data)

        total_reach = safe_sum(ym_data, 'Total Post Reach')
        total_likes = safe_sum(ym_data, 'Likes')
        total_comments = safe_sum(ym_data, 'Comments')
        total_shares = safe_sum(ym_data, 'Shares')
        total_saves = safe_sum(ym_data, 'Saves')
        total_interaction = total_likes + total_comments + total_shares + total_saves

        total_stories = 0
        total_story_reach = 0
        total_story_shares = 0
        total_story_clicks = 0
        if df_ig_story is not None and not df_ig_story.empty and '_Period' in df_ig_story.columns:
            story_data = df_ig_story[df_ig_story['_Period'] == ym]
            total_stories = len(story_data)
            total_story_reach = safe_sum(story_data, 'Total Reach')
            total_story_shares = safe_sum(story_data, 'Shares')
            total_story_clicks = safe_sum(story_data, 'Link clicks')

        row_data = {
            'Feeds Posted': total_posts,
            'Total Post Likes': int(total_likes),
            'Total Post Comments': int(total_comments),
            'Total Post Shares': int(total_shares),
            'Total Post Saves': int(total_saves),
            'Total Post Interaction': int(total_interaction),
            'Total Post Reach': int(total_reach),
            'Average Post Likes': round(total_likes / total_posts, 1) if total_posts > 0 else 0,
            'Average Post Comments': round(total_comments / total_posts, 1) if total_posts > 0 else 0,
            'Average Post Shares': round(total_shares / total_posts, 1) if total_posts > 0 else 0,
            'Average Post Saves': round(total_saves / total_posts, 1) if total_posts > 0 else 0,
            'Average Post Interaction': round(total_interaction / total_posts, 1) if total_posts > 0 else 0,
            'Average Post Reach': round(total_reach / total_posts, 1) if total_posts > 0 else 0,
            'Stories Posted': total_stories,
            'Total Stories Link Clicks': int(total_story_clicks),
            'Total Stories Shares': int(total_story_shares),
            'Total Stories Reach': int(total_story_reach),
            'Average Stories Link Clicks': round(total_story_clicks / total_stories, 1) if total_stories > 0 else None,
            'Average Stories Shares': round(total_story_shares / total_stories, 1) if total_stories > 0 else None,
            'Average Stories Reach': round(total_story_reach / total_stories, 1) if total_stories > 0 else None,
        }

        for metric in metric_order:
            metrics_list.append({'Metric': metric, ym: row_data[metric]})

    if not metrics_list:
        return pd.DataFrame()

    df_metrics = pd.DataFrame(metrics_list)
    df_metrics = df_metrics.groupby('Metric', sort=False).first().reset_index()
    return df_metrics


def generate_ig_key_metrics_cm_all(df_ig_post, df_ig_story):
    logging.info("Generating IG Key Metric (CM) (all periods) ...")

    if df_ig_post is None or df_ig_post.empty or '_Period' not in df_ig_post.columns:
        return pd.DataFrame()

    if 'Pillar' in df_ig_post.columns:
        mask = ~df_ig_post['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_ig_post = df_ig_post[mask].copy()
    if df_ig_story is not None and not df_ig_story.empty and 'Pillar' in df_ig_story.columns:
        mask = ~df_ig_story['Pillar'].astype(str).str.strip().str.lower().isin(['ecommerce', 'gnc'])
        df_ig_story = df_ig_story[mask].copy()

    return generate_ig_key_metrics_all(df_ig_post, df_ig_story)


# ==============================================================================
# FOLLOWERS FROM API LOG (for Key Metrics tabs)
# ==============================================================================
def fetch_followers_monthly(worksheet_name):
    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
        ws = sh.worksheet(worksheet_name)
        df = pd.DataFrame(ws.get_all_records())
        if df.empty or "Date" not in df.columns:
            return {}

        df["_dt"] = pd.to_datetime(df["Date"].astype(str).str.strip(),
                                   format="%d/%m/%Y", errors="coerce")
        df = df.dropna(subset=["_dt"]).sort_values("_dt")

        if "Total Followers" in df.columns:
            df["Total Followers"] = pd.to_numeric(
                df["Total Followers"], errors="coerce"
            ).ffill().bfill().astype(int)
        if "Followers Net" in df.columns:
            df["Followers Net"] = pd.to_numeric(
                df["Followers Net"], errors="coerce"
            ).fillna(0).astype(int)

        result = {}
        for (y, m), group in df.groupby([df["_dt"].dt.year, df["_dt"].dt.month]):
            period = f"{y}-{m:02d}"
            first_row = group.iloc[0]
            total = int(first_row["Total Followers"]) if "Total Followers" in df.columns else 0
            net = int(group["Followers Net"].sum()) if "Followers Net" in df.columns else 0
            result[period] = {"total": total, "net": net}

        logging.info(f"  {worksheet_name}: {len(result)} periods of follower data")
        return result
    except Exception as e:
        logging.error(f"Failed to fetch followers from {worksheet_name}: {e}")
        return {}


def add_followers_to_key_metrics(df_metrics, followers_map):
    if not followers_map or df_metrics.empty:
        return df_metrics

    period_cols = [c for c in df_metrics.columns if c != "Metric"]
    tf_row = {"Metric": "Total Followers"}
    ng_row = {"Metric": "Net Followers Growth"}
    for p in period_cols:
        if p in followers_map:
            tf_row[p] = followers_map[p]["total"]
            ng_row[p] = followers_map[p]["net"]
        else:
            tf_row[p] = None
            ng_row[p] = None

    new_rows = pd.DataFrame([tf_row, ng_row])
    return pd.concat([new_rows, df_metrics], ignore_index=True)


# ==============================================================================
# API-SOURCED TABS (from 3b)
# ==============================================================================
def build_api_tabs():
    logging.info("Building API tabs from Google Sheet API Logs ...")
    result = {}

    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
    except Exception as e:
        logging.error(f"Failed to connect for API logs: {e}")
        return result

    # ── Unique Page View + FB Followers (from FB API Log) ──
    df_fb = pd.DataFrame()
    try:
        ws = sh.worksheet("FB API Log")
        df_fb = pd.DataFrame(ws.get_all_records())
        if not df_fb.empty and "Date" in df_fb.columns:
            df_view = df_fb[["Date", "Unique Page View"]].copy()
            df_view["Unique Page View"] = pd.to_numeric(
                df_view["Unique Page View"], errors="coerce"
            ).fillna(0).astype(int)
            result["Unique Page View"] = df_view
            logging.info(f"  Unique Page View: {len(df_view)} rows")
    except Exception as e:
        logging.error(f"  FB API Log error: {e}")

    if not df_fb.empty and "Date" in df_fb.columns:
        cols = ["Date", "Total Followers", "Followers Gain", "Followers Loss", "Followers Net"]
        available = [c for c in cols if c in df_fb.columns]
        if available:
            df_followers = df_fb[available].copy()
            for c in ["Total Followers", "Followers Gain", "Followers Loss", "Followers Net"]:
                if c in df_followers.columns:
                    df_followers[c] = pd.to_numeric(
                        df_followers[c], errors="coerce"
                    ).fillna(0).astype(int)
            result["FB Followers"] = df_followers
            logging.info(f"  FB Followers: {len(df_followers)} rows")

    # ── IG Followers (from IG API Log, with forward-fill) ──
    try:
        ws_ig = sh.worksheet("IG API Log")
        df_ig = pd.DataFrame(ws_ig.get_all_records())
        if not df_ig.empty and "Date" in df_ig.columns:
            df_ig["Date"] = df_ig["Date"].astype(str).str.strip()
            df_ig["Followers Net"] = pd.to_numeric(
                df_ig.get("Followers Net", 0), errors="coerce"
            ).fillna(0).astype(int)

            dates = pd.to_datetime(df_ig["Date"], format="%d/%m/%Y", errors="coerce")
            df_ig["_dt"] = dates

            if len(df_ig) > 0:
                min_dt = df_ig["_dt"].min()
                max_dt = df_ig["_dt"].max()
                full_range = pd.date_range(min_dt, max_dt, freq="D")
                df_daily = pd.DataFrame({"_dt": full_range})
                df_daily = df_daily.merge(df_ig, on="_dt", how="left")

                if "Total Followers" in df_daily.columns:
                    tf = pd.to_numeric(df_daily["Total Followers"], errors="coerce")
                    df_daily["Total Followers"] = tf.ffill().bfill().astype(int)

                df_daily["Followers Net"] = df_daily["Followers Net"].fillna(0).astype(int)

                mask_empty = df_daily["Followers Net"] == 0
                if mask_empty.any() and "Total Followers" in df_daily.columns:
                    diffs = df_daily["Total Followers"].diff().fillna(0).astype(int)
                    df_daily.loc[mask_empty, "Followers Net"] = diffs[mask_empty]

                df_out = pd.DataFrame({
                    "Date": df_daily["_dt"].dt.strftime("%d/%m/%Y"),
                    "Total Followers": df_daily["Total Followers"].astype(int),
                    "Followers Net": df_daily["Followers Net"].astype(int),
                })
                result["IG Followers"] = df_out
                logging.info(f"  IG Followers: {len(df_out)} rows (forward-filled)")
    except Exception as e:
        logging.error(f"  IG API Log error: {e}")

    return result


def build_reach_funnel_tab():
    logging.info("Building Reach Funnel tab ...")
    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_KEY_METRICS_ID)
        ws = sh.worksheet("All FB Posts_New Cat")
        records = ws.get_all_records()
        df = pd.DataFrame(records)
    except Exception as e:
        logging.error(f"Failed to read All FB Posts_New Cat: {e}")
        return None

    if df.empty:
        return None

    col_map = {}
    for c in df.columns:
        stripped = c.strip()
        if stripped in ("Publish time", "Month No.", "Day", "Organic reach", "Paid reach"):
            col_map[c] = stripped

    mn_col = next((c for c, v in col_map.items() if v == "Month No."), None)
    day_col = next((c for c, v in col_map.items() if v == "Day"), None)
    pt_col = next((c for c, v in col_map.items() if v == "Publish time"), None)

    df_out = pd.DataFrame(index=df.index)
    df_out["Year"] = 2026
    if mn_col:
        df_out["Month No."] = pd.to_numeric(df[mn_col], errors="coerce").astype("Int64")
    if day_col:
        df_out["Day"] = pd.to_numeric(df[day_col], errors="coerce").astype("Int64")
    if pt_col:
        df_out["Publish time"] = df[pt_col].astype(str).str.strip()
    else:
        df_out["Publish time"] = ""

    for label in ("Organic reach", "Paid reach"):
        src_col = next((c for c, v in col_map.items() if v == label), None)
        if src_col:
            df_out[label] = pd.to_numeric(df[src_col].astype(str).str.replace(",", ""), errors="coerce").fillna(0).astype(int)
        else:
            df_out[label] = 0

    df_out = df_out.dropna(subset=["Year", "Month No."]).reset_index(drop=True)
    logging.info(f"  Reach Funnel: {len(df_out)} rows")
    return df_out


# ==============================================================================
# LINKEDIN PERFORMANCE CORE GENERATORS (UNFILTERED ALL-PERIOD ENGINE)
# ==============================================================================
def build_linkedin_posts_perf():
    logging.info("Building LinkedIn Posts Perf tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()

    files = list(LINKEDIN_RAW_DIR.glob("*manningshk_content*"))
    all_dfs = []
    for f in files:
        try:
            df = smart_read_linkedin_excel(f, sheet_keyword='all posts') if f.suffix in ['.xlsx', '.xls'] else pd.read_csv(f, header=None)
            if f.suffix not in ['.xlsx', '.xls']:
                header_row_idx = 0
                for i, row in df.iterrows():
                    if len(row.dropna()) > 1:
                        header_row_idx = i
                        break
                df.columns = df.iloc[header_row_idx]
                df = df.iloc[header_row_idx+1:].reset_index(drop=True)
                df.columns = df.columns.str.strip()
            
            if not df.empty: all_dfs.append(df)
        except Exception as e: logging.error(f"  Error reading file {f.name}: {e}")
    if not all_dfs: return pd.DataFrame()

    df_all = pd.concat(all_dfs, ignore_index=True).drop_duplicates()
    date_col = next((c for c in df_all.columns if str(c).lower().strip() in ['created date', 'date']), None)
    if not date_col: return pd.DataFrame()

    fmt_dates, months, years, days = process_linkedin_dates(df_all[date_col], col_name=f"Content Posts: {date_col}")
    df_all['_fmt_date'], df_all['_m'], df_all['_y'], df_all['_d'] = fmt_dates, months, years, days
    
    df_filtered = df_all[df_all['_y'].notna() & df_all['_m'].notna()].copy()
    if df_filtered.empty: return pd.DataFrame()

    # Chronological Sort Ascending
    df_filtered = df_filtered.sort_values(by=['_y', '_m', '_d'], ascending=True).copy()

    def get_col_or_zero(df, keywords):
        c = next((col for col in df.columns if any(kw in str(col).lower() for kw in keywords)), None)
        return pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int) if c else pd.Series(0, index=df.index)

    df_out = pd.DataFrame()
    df_out['Month'] = df_filtered['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
    df_out['Post date'] = df_filtered['_fmt_date']
    df_out['Post Link'] = df_filtered[next((c for c in df_filtered.columns if any(k in str(c).lower() for k in ['link', 'url'])), df_filtered.columns[0])].astype(str).str.strip()
    df_out['Post message'] = df_filtered[next((c for c in df_filtered.columns if any(kw in str(c).lower() for kw in ['title', 'message', 'description'])), df_filtered.columns[0])].astype(str).str.strip()
    df_out['Impressions'] = get_col_or_zero(df_filtered, ['impressions'])

    ct_col = next((c for c in df_filtered.columns if 'content type' in str(c).lower()), None)
    v_col = next((c for c in df_filtered.columns if 'views' in str(c).lower()), None)
    df_out['Video Views'] = df_filtered.apply(lambda r: pd.to_numeric(r[v_col], errors='coerce') if ct_col and v_col and 'video' in str(r[ct_col]).lower() else 0, axis=1).fillna(0).astype(int)

    df_out['Clicks'] = get_col_or_zero(df_filtered, ['clicks'])
    df_out['Likes'] = get_col_or_zero(df_filtered, ['likes', 'reactions'])
    df_out['Comments'] = get_col_or_zero(df_filtered, ['comments'])
    df_out['Reposts'] = get_col_or_zero(df_filtered, ['reposts', 'shares'])
    df_out['Follows'] = get_col_or_zero(df_filtered, ['follows'])
    df_out['Total Interactions'] = df_out['Likes'] + df_out['Comments'] + df_out['Reposts'] + df_out['Follows']
    return df_out


def build_linkedin_page_perf():
    logging.info("Building LinkedIn Page Perf tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()

    df_metrics = pd.DataFrame()
    for f in LINKEDIN_RAW_DIR.glob("*manningshk_content*"):
        if f.suffix in ['.xlsx', '.xls']:
            s = smart_read_linkedin_excel(f, sheet_keyword='metrics')
            if not s.empty: df_metrics = pd.concat([df_metrics, s], ignore_index=True)

    df_visitors = pd.DataFrame()
    for f in LINKEDIN_RAW_DIR.glob("*manningshk_visitors*"):
        if f.suffix in ['.xlsx', '.xls']:
            s = smart_read_linkedin_excel(f, sheet_keyword='visitor')
            if not s.empty: df_visitors = pd.concat([df_visitors, s], ignore_index=True)

    if df_metrics.empty or df_visitors.empty: return pd.DataFrame()

    m_date_col = next((c for c in df_metrics.columns if 'date' in str(c).lower()), df_metrics.columns[0])
    v_date_col = next((c for c in df_visitors.columns if 'date' in str(c).lower()), df_visitors.columns[0])

    m_fmt, m_m, m_y, m_d = process_linkedin_dates(df_metrics[m_date_col], col_name=f"Metrics Base: {m_date_col}")
    df_metrics['_fmt_date'], df_metrics['_m'], df_metrics['_y'], df_metrics['_d'] = m_fmt, m_m, m_y, m_d
    v_fmt, v_m, v_y, v_d = process_linkedin_dates(df_visitors[v_date_col], col_name=f"Visitors Base: {v_date_col}")
    df_visitors['_fmt_date'], df_visitors['_m'], df_visitors['_y'], df_visitors['_d'] = v_fmt, v_m, v_y, v_d

    df_metrics = df_metrics[df_metrics['_y'].notna()].drop_duplicates(subset=['_fmt_date'])
    df_visitors = df_visitors[df_visitors['_y'].notna()].drop_duplicates(subset=['_fmt_date'])
    if df_metrics.empty or df_visitors.empty: return pd.DataFrame()

    df_merged = pd.merge(df_metrics, df_visitors, on='_fmt_date', how='outer', suffixes=('', '_v'))
    if '_m' not in df_merged.columns or df_merged['_m'].isnull().all(): df_merged['_m'] = df_merged['_m_v']
    if '_y' not in df_merged.columns or df_merged['_y'].isnull().all(): df_merged['_y'] = df_merged['_y_v']
    if '_d' not in df_merged.columns or df_merged['_d'].isnull().all(): df_merged['_d'] = df_merged['_d_v']

    # Chronological Sort Ascending
    df_merged = df_merged.sort_values(by=['_y', '_m', '_d'], ascending=True).copy()

    def get_val(df, keywords):
        c = next((col for col in df.columns if any(kw in str(col).lower() for kw in keywords)), None)
        return pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int) if c else pd.Series(0, index=df.index)

    df_out = pd.DataFrame()
    df_out['Month'] = df_merged['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
    df_out['Date'] = df_merged['_fmt_date']
    df_out['Desktop Views'] = get_val(df_merged, ['page views (desktop)'])
    df_out['Mobile Views'] = get_val(df_merged, ['page views (mobile)'])
    df_out['Total Page Views'] = get_val(df_merged, ['page views (total)'])
    df_out['Likes'] = get_val(df_merged, ['reactions (total)'])
    df_out['Comments'] = get_val(df_merged, ['comments (total)'])
    df_out['Shares'] = get_val(df_merged, ['reposts (total)', 'shares (total)'])
    df_out['Clicks'] = get_val(df_merged, ['clicks(total)', 'clicks (total)'])
    df_out['Impressions'] = get_val(df_merged, ['impressions (total)'])
    df_out['Likes (Organic)'] = get_val(df_merged, ['reactions (organic)'])
    df_out['Comments (Organic)'] = get_val(df_merged, ['comments (organic)'])
    df_out['Shares (Organic)'] = get_val(df_merged, ['reposts (organic)', 'shares (organic)'])
    df_out['Clicks (Organic)'] = get_val(df_merged, ['clicks(organic)', 'clicks (organic)'])
    df_out['Impressions (Organic)'] = get_val(df_merged, ['impressions (organic)'])
    df_out['Likes (Paid)'] = get_val(df_merged, ['reactions (sponsored)'])
    df_out['Comments (Paid)'] = get_val(df_merged, ['comments (sponsored)'])
    df_out['Shares (Paid)'] = get_val(df_merged, ['reposts (sponsored)', 'shares (sponsored)'])
    df_out['Clicks (Paid)'] = get_val(df_merged, ['clicks(sponsored)', 'clicks (sponsored)'])
    df_out['Impressions (Paid)'] = get_val(df_merged, ['impressions (sponsored)'])
    return df_out


def build_linkedin_follower_log():
    logging.info("Building LinkedIn Follower Log tab ...")
    if not LINKEDIN_RAW_DIR.exists(): return pd.DataFrame()
    log_date_str = pd.Timestamp.now().strftime('%m/%d/%Y')

    all_frames = []
    net_rolling = 0

    # 1. Historical Slices (Jan - Jun)
    f_log = LINKEDIN_RAW_DIR / "Follower log_202601_06.xlsx"
    if f_log.exists():
        try:
            df_log = smart_read_linkedin_excel(f_log)
            d_col = next((c for c in df_log.columns if 'date' in str(c).lower()), df_log.columns[0])
            fmt, mm, yy, dd = process_linkedin_dates(df_log[d_col], col_name=f"Hist Follower Log: {d_col}")
            df_log['_fmt_date'], df_log['_m'], df_log['_y'], df_log['_d'] = fmt, mm, yy, dd
            df_hist = df_log[df_log['_y'].notna()].copy()
            if not df_hist.empty:
                df_out_hist = pd.DataFrame()
                df_out_hist['Month'] = df_hist['_m'].apply(lambda x: MONTH_MAP[int(x)] if pd.notna(x) else "")
                df_out_hist['Date Polled'] = df_hist['_fmt_date']
                for col in ['Category', 'Demographic', 'Total Followers']:
                    match = next((c for c in df_hist.columns if str(c).lower() == col.lower()), None)
                    df_out_hist[col] = df_hist[match] if match else ""
                df_out_hist['Log Date'] = log_date_str
                df_out_hist['_y'], df_out_hist['_m'], df_out_hist['_d'] = df_hist['_y'], df_hist['_m'], df_hist['_d']
                all_frames.append(df_out_hist)
                
                # Carry forward Net follower totals from last day of log
                df_net_rows = df_out_hist[df_out_hist['Category'].astype(str).str.strip().str.lower() == 'net']
                if not df_net_rows.empty:
                    df_net_rows = df_net_rows.sort_values(by=['_y', '_m', '_d'])
                    last_net = pd.to_numeric(df_net_rows.iloc[-1]['Total Followers'], errors='coerce')
                    if not pd.isna(last_net):
                        net_rolling = int(last_net)
        except Exception as e: logging.error(f"  Error slicing historical file log: {e}")

    # 2. Daily logs (July onwards)
    f_files = list(LINKEDIN_RAW_DIR.glob("*manningshk_followers*"))
    
    daily_raw_rows = []
    demo_tasks = []
    
    for f in f_files:
        try:
            xl = pd.ExcelFile(f)
            sheet = next((s for s in xl.sheet_names if 'new followers' in s.lower() or s.lower() == 'followers'), None)
            if not sheet:
                sheet = next((s for s in xl.sheet_names if 'followers' in s.lower() and 'demographic' not in s.lower()), None)
            if not sheet: continue
            
            df_nf = smart_read_linkedin_excel(xl, sheet_keyword=sheet)
            if df_nf.empty: continue
            
            d_col = next((c for c in df_nf.columns if 'date' in str(c).lower()), df_nf.columns[0])
            fmt, mm, yy, dd = process_linkedin_dates(df_nf[d_col], col_name=f"Daily Followers Log ({f.name}): {d_col}")
            df_nf['_fmt_date'], df_nf['_m'], df_nf['_y'], df_nf['_d'] = fmt, mm, yy, dd
            
            # STRICT GUARDRAIL: Restrict the daily log data ingestion exclusively to July onwards (>= 7)
            # This perfectly prevents any duplicate data pollution between January and June.
            df_nf_filtered = df_nf[(df_nf['_y'] == 2026) & (df_nf['_m'] >= 7)].copy()
            if df_nf_filtered.empty: continue
            
            org_c = next((c for c in df_nf_filtered.columns if 'organic' in str(c).lower()), None)
            paid_c = next((c for c in df_nf_filtered.columns if 'paid' in str(c).lower() or 'sponsored' in str(c).lower()), None)
            
            for _, r in df_nf_filtered.iterrows():
                org_v = pd.to_numeric(r[org_c], errors='coerce') if org_c else 0
                paid_v = pd.to_numeric(r[paid_c], errors='coerce') if paid_c else 0
                daily_raw_rows.append({
                    '_fmt_date': r['_fmt_date'], '_m': r['_m'], '_y': r['_y'], '_d': r['_d'],
                    'organic': int(org_v) if not pd.isna(org_v) else 0,
                    'paid': int(paid_v) if not pd.isna(paid_v) else 0
                })
                
            # Isolate the exact demographical end-state window matching the last date string record inside the file
            last_rec = df_nf_filtered.sort_values(by=['_y', '_m', '_d']).iloc[-1]
            demo_tasks.append({
                'xl_obj': xl,
                'last_date_str': last_rec['_fmt_date'],
                '_m': last_rec['_m'],
                '_y': last_rec['_y'],
                '_d': last_rec['_d']
            })
        except Exception as e:
            logging.error(f"  Error reading follower file {f.name}: {e}")

    # Build unique chronological daily tracking frame
    if daily_raw_rows:
        df_daily_all = pd.DataFrame(daily_raw_rows).drop_duplicates(subset=['_fmt_date'])
        df_daily_all = df_daily_all.sort_values(by=['_y', '_m', '_d']).reset_index(drop=True)
        
        daily_output_rows = []
        for _, r in df_daily_all.iterrows():
            m_str = MONTH_MAP[int(r['_m'])] if pd.notna(r['_m']) else ""
            org_v = r['organic']
            paid_v = r['paid']
            net_rolling += (org_v + paid_v)
            
            daily_output_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Organic', 'Demographic': '', 'Total Followers': org_v, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})
            daily_output_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Paid', 'Demographic': '', 'Total Followers': paid_v, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})
            daily_output_rows.append({'Month': m_str, 'Date Polled': r['_fmt_date'], 'Category': 'Net', 'Demographic': '', 'Total Followers': net_rolling, 'Log Date': log_date_str, '_y': r['_y'], '_m': r['_m'], '_d': r['_d']})
            
        # Append top 5 breakdown profiles tied to the specific window's target date
        for task in demo_tasks:
            xl = task['xl_obj']
            m_str = MONTH_MAP[int(task['_m'])] if pd.notna(task['_m']) else ""
            demo_categories = {'Location': ['location'], 'Job function': ['function'], 'Seniority': ['seniority'], 'Industry': ['industry'], 'Company size': ['size']}
            
            for cat_name, keywords in demo_categories.items():
                s_name = next((s for s in xl.sheet_names if any(kw in s.lower() for kw in keywords)), None)
                if s_name:
                    df_demo = smart_read_linkedin_excel(xl, sheet_keyword=s_name)
                    if df_demo.empty: continue
                    
                    cnt_col = next((c for c in df_demo.columns if any(kw in str(c).lower() for kw in ['follower', 'count', 'total'])), df_demo.columns[1])
                    df_demo[cnt_col] = pd.to_numeric(df_demo[cnt_col], errors='coerce').fillna(0)
                    
                    for _, d_row in df_demo.sort_values(by=cnt_col, ascending=False).head(5).iterrows():
                        daily_output_rows.append({
                            'Month': m_str, 'Date Polled': task['last_date_str'],
                            'Category': cat_name, 'Demographic': str(d_row[df_demo.columns[0]]), 'Total Followers': int(d_row[cnt_col]), 'Log Date': log_date_str,
                            '_y': task['_y'], '_m': task['_m'], '_d': task['_d']
                        })
                        
        if daily_output_rows:
            all_frames.append(pd.DataFrame(daily_output_rows))

    if all_frames:
        df_res = pd.concat(all_frames, ignore_index=True)
        df_res = df_res.sort_values(by=['_y', '_m', '_d'], ascending=True).drop(columns=['_y', '_m', '_d']).reset_index(drop=True)
        return df_res
    return pd.DataFrame()

# ==============================================================================
# CACHE AND OUTPUT CONFIGURATION
# ==============================================================================
output_frames = {}

EXPECTED_TABS = [
    "FB Pivot (Category)",
    "FB Pivot (Post Type)",
    "FB Pivot (Pillar)",
    "FB Key Metrics",
    "FB Key Metric (CM)",
    "FB Wall Post Performance",
    "Category Performance - BAU",
    "Sub-Category Performance - PNP",
    "Pillar Performance - CRM",
    "Pillar Performance - Ecommerce",
    "Pillar Performance - GNC",
    "Pillar Performance - Branding",
    "Pillar Performance - Category",
    "Pillar Performance - Sales",
    "Pillar Performance - Others",
    "IG Story Performance",
    "IG Story Pivot",
    "IG Wall Post Performance",
    "IG Pivot (Pillar)",
    "IG Key Metrics",
    "IG Key Metric (CM)",
    "Master Comments Base",
    "Sentiment Summary (Category)",
    "Sentiment Summary (Type)",
    "LinkedIn Posts Perf",
    "LinkedIn Page Perf",
    "LinkedIn Follower Log"
]

def write_to_sheet(tab_name, dataframe):
    if dataframe is None:
        dataframe = pd.DataFrame()
    output_frames[tab_name] = dataframe.copy()
    if dataframe.empty:
        logging.info(f"  Tab '{tab_name}': empty (placeholder will be created)")
    else:
        logging.info(f"  Cached tab '{tab_name}': {len(dataframe)} rows")


# ==============================================================================
# MAIN PIPELINE
# ==============================================================================
def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info("=== 3_new_v2: Master Aggregation (ALL periods) ===")

    try:
        client = get_gspread_client()
        sh = client.open_by_key(GSHEET_TRACKER_ID)
        logging.info("Connected to Social Tracker Google Sheet.")
    except Exception as e:
        logging.error(f"Failed to connect: {e}")
        return

    # ------------------------------------------------------------------
    # 0. FETCH FOLLOWER DATA FROM API LOGS
    # ------------------------------------------------------------------
    fb_followers_map = fetch_followers_monthly("FB API Log")
    ig_followers_map = fetch_followers_monthly("IG API Log")

    # ------------------------------------------------------------------
    # 1. FACEBOOK DATA
    # ------------------------------------------------------------------
    logging.info("Processing FB data (All FB Posts_New Cat) ...")
    try:
        fb_ws = sh.worksheet("All FB Posts_New Cat")
        df_fb = pd.DataFrame(fb_ws.get_all_records())
        if not df_fb.empty:
            df_fb = normalize_columns(df_fb)
            df_fb_filtered = assign_periods(df_fb, 'Publish time')

            if len(df_fb_filtered) > 0:
                try:
                    df_fb_key = generate_fb_key_metrics_all(df_fb_filtered.copy())
                    if not df_fb_key.empty:
                        df_fb_key = add_followers_to_key_metrics(df_fb_key, fb_followers_map)
                        write_to_sheet("FB Key Metrics", df_fb_key)
                except Exception as e:
                    logging.error(f"FB Key Metrics error: {e}")

                try:
                    df_fb_key_cm = generate_fb_key_metrics_cm_all(df_fb_filtered.copy())
                    if not df_fb_key_cm.empty:
                        df_fb_key_cm = add_followers_to_key_metrics(df_fb_key_cm, fb_followers_map)
                        write_to_sheet("FB Key Metric (CM)", df_fb_key_cm)
                except Exception as e:
                    logging.error(f"FB Key Metric (CM) error: {e}")

                if 'Dark Post' in df_fb_filtered.columns:
                    df_fb_filtered = df_fb_filtered[df_fb_filtered['Dark Post'] != 'Y']
                    logging.info(f"  Excluded dark posts: {len(df_fb_filtered)} rows remaining")
                if 'Month' in df_fb_filtered.columns:
                    df_fb_filtered = df_fb_filtered.drop(columns=['Month'])

                if 'Post type' in df_fb_filtered.columns:
                    df_fb_filtered['Post type'] = df_fb_filtered['Post type'].astype(str).str.strip().replace({
                        'Photo': 'Photos', 'Single Photo': 'Photos',
                        'Video': 'Videos', 'Link': 'Links',
                    })

                df_fb_filtered['Interactions_Numeric'] = clean_numeric_col(df_fb_filtered, 'Interactions')
                fb_pillar_col = get_actual_col_name(df_fb_filtered, ['Pillar'])
                fb_category_col = get_actual_col_name(df_fb_filtered, ['Category', 'category'])
                fb_post_type_col = get_actual_col_name(df_fb_filtered, ['Post type', 'post type', 'Post Type'])
                fb_subcat_col = get_actual_col_name(df_fb_filtered, ['Sub-Category', 'sub-category', 'Sub Category'])

                pivot_fb_category = df_fb_filtered.groupby(['_Period', fb_category_col]).agg(
                    No_of_post=(fb_category_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_category_col: 'Category',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                pivot_fb_post_type = df_fb_filtered.groupby(['_Period', fb_post_type_col]).agg(
                    No_of_post=(fb_post_type_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_post_type_col: 'Post type',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                pivot_fb_pillar = df_fb_filtered.groupby(['_Period', fb_pillar_col]).agg(
                    No_of_post=(fb_pillar_col, 'count'),
                    Total_Interactions=('Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', fb_pillar_col: 'Pillar',
                    'No_of_post': 'No. of Posts', 'Total_Interactions': 'Total Interactions'})

                write_to_sheet("FB Pivot (Category)", pivot_fb_category)
                write_to_sheet("FB Pivot (Post Type)", pivot_fb_post_type)
                write_to_sheet("FB Pivot (Pillar)", pivot_fb_pillar)

                df_fb_filtered = df_fb_filtered.drop(columns=['Interactions_Numeric'])
                df_fb_filtered.columns = df_fb_filtered.columns.str.strip()

                fb_perf_cols = [
                    '_Period', '_Year', '_Month', '_Day',
                    'Permalink', 'Description', 'Post type', 'Pillar', 'Category',
                    'Sub-Category', 'Campaign Name', 'Publish time', 'Reactions',
                    'Comments', 'Shares', 'Interactions', 'Reach', 'Views',
                    'Link Clicks', 'Photo Clicks', '3-second video views',
                    'Video Length', 'Organic reach', 'Paid reach', 'Dark Post',
                    'Organic', 'Paid', 'Month No.', 'Day', 'Hour', 'Average Watch Time'
                ]
                fb_numeric_cols = [
                    'Reactions', 'Comments', 'Shares', 'Interactions', 'Reach', 'Views',
                    'Link Clicks', 'Photo Clicks', '3-second video views', 'Organic reach',
                    'Paid reach', 'Month No.', 'Day', 'Hour', 'Video Length', 'Average Watch Time'
                ]

                def prep_fb_perf(df_subset):
                    if df_subset is None or len(df_subset) == 0:
                        return pd.DataFrame(columns=fb_perf_cols)
                    result = df_subset.copy()
                    for col in fb_perf_cols:
                        if col not in result.columns:
                            result[col] = ""
                    result = result[fb_perf_cols]
                    for col in fb_numeric_cols:
                        if col in result.columns:
                            result[col] = pd.to_numeric(result[col].astype(str).str.replace(',', ''), errors='coerce')
                    return result

                write_to_sheet("FB Wall Post Performance", prep_fb_perf(df_fb_filtered))

                perf_breakdown_cols = ['_Period', '_Year', '_Month', '_Day', 'Permalink', 'Description', 'Post type', 'Pillar', 'Category', 'Sub-Category',
                                       'Campaign Name', 'Publish time', 'Reactions', 'Comments', 'Shares',
                                       'Interactions', 'Reach', 'Views', 'Link Clicks', 'Photo Clicks']

                def prep_breakdown(df_subset):
                    df_out = prep_fb_perf(df_subset)
                    cols = [c for c in perf_breakdown_cols if c in df_out.columns]
                    return df_out[cols] if cols else df_out

                cat_clean = df_fb_filtered['Category'].astype(str).str.lower().str.strip() if 'Category' in df_fb_filtered.columns else pd.Series(index=df_fb_filtered.index).fillna("")
                subcat_clean = df_fb_filtered[fb_subcat_col].astype(str).str.lower().str.strip() if fb_subcat_col in df_fb_filtered.columns else pd.Series(index=df_fb_filtered.index).fillna("")
                write_to_sheet("Category Performance - BAU", prep_breakdown(df_fb_filtered[cat_clean == 'bau promotion']))
                write_to_sheet("Sub-Category Performance - PNP", prep_breakdown(df_fb_filtered[subcat_clean == 'friday pnp']))
                write_to_sheet("Pillar Performance - CRM", prep_breakdown(df_fb_filtered[cat_clean.str.contains('yuu', na=False) | (cat_clean == 'weekly bau')]))
                write_to_sheet("Pillar Performance - Ecommerce", prep_breakdown(df_fb_filtered[cat_clean.str.contains('ecom', na=False)]))
                write_to_sheet("Pillar Performance - GNC", prep_breakdown(df_fb_filtered[cat_clean == 'gnc']))
                pillar_clean = df_fb_filtered['Pillar'].astype(str).str.strip() if 'Pillar' in df_fb_filtered.columns else pd.Series(index=df_fb_filtered.index).fillna("")
                write_to_sheet("Pillar Performance - Branding", prep_breakdown(df_fb_filtered[pillar_clean == 'Branding']))
                write_to_sheet("Pillar Performance - Category", prep_breakdown(df_fb_filtered[pillar_clean == 'Category']))
                write_to_sheet("Pillar Performance - Sales", prep_breakdown(df_fb_filtered[pillar_clean == 'Sales']))
                write_to_sheet("Pillar Performance - Others", prep_breakdown(df_fb_filtered[cat_clean == 'others']))
    except Exception as e:
        logging.error(f"FB processing error: {e}")

    # ------------------------------------------------------------------
    # 2. IG STORY DATA
    # ------------------------------------------------------------------
    logging.info("Processing IG Story data (All_IGS) ...")
    df_igs_for_metrics = pd.DataFrame()
    try:
        igs_ws = sh.worksheet("All_IGS")
        df_igs = pd.DataFrame(igs_ws.get_all_records())
        if not df_igs.empty:
            df_igs = normalize_columns(df_igs)
            df_igs_filtered = assign_periods(df_igs, 'Publish time')
            if len(df_igs_filtered) > 0:
                igs_pillar_col = get_actual_col_name(df_igs_filtered, ['Pillar'])
                link_click_col = get_actual_col_name(df_igs_filtered, ['Link clicks', 'link clicks'])
                df_igs_filtered['Link_Clicks_Numeric'] = clean_numeric_col(df_igs_filtered, link_click_col)

                pivot_igs_pillar = df_igs_filtered.groupby(['_Period', igs_pillar_col]).agg(
                    No_of_post=(igs_pillar_col, 'count'),
                    Total_Link_Clicks=('Link_Clicks_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', igs_pillar_col: 'Pillar',
                    'No_of_post': 'No. of Posts', 'Total_Link_Clicks': 'SUM of Link clicks'})
                df_igs_filtered = df_igs_filtered.drop(columns=['Link_Clicks_Numeric'])

                df_igs_for_metrics = df_igs_filtered.copy()

                igs_cols_to_keep = ['_Period', '_Year', '_Month', '_Day', 'Publish time', 'Description', 'Pillar', 'Category', 'Total Reach', 'Shares', 'Link clicks']
                igs_cols_actual = [get_actual_col_name(df_igs_filtered, [c, c.lower(), c.capitalize()]) for c in igs_cols_to_keep]
                df_igs_filtered = df_igs_filtered[[c for c in igs_cols_actual if c in df_igs_filtered.columns]]
                for col in ['Total Reach', 'Shares', 'Link clicks']:
                    actual = get_actual_col_name(df_igs_filtered, [col, col.lower()])
                    if actual in df_igs_filtered.columns:
                        df_igs_filtered[actual] = pd.to_numeric(
                            df_igs_filtered[actual].astype(str).str.replace(',', ''), errors='coerce')

                write_to_sheet("IG Story Performance", df_igs_filtered)
                write_to_sheet("IG Story Pivot", pivot_igs_pillar)
    except Exception as e:
        logging.error(f"IG Story error: {e}")

    # ------------------------------------------------------------------
    # 3. IG WALL POST DATA
    # ------------------------------------------------------------------
    logging.info("Processing IG Wall Post data (All_IG_Posts) ...")
    df_igp_for_metrics = pd.DataFrame()
    try:
        igp_ws = sh.worksheet("All_IG_Posts")
        df_igp = pd.DataFrame(igp_ws.get_all_records())
        if not df_igp.empty:
            df_igp = normalize_columns(df_igp)
            df_igp_filtered = assign_periods(df_igp, 'Publish time')
            if len(df_igp_filtered) > 0:
                pub_time_col = get_actual_col_name(df_igp_filtered, ['Publish time', 'publish time'])
                views_col = get_actual_col_name(df_igp_filtered, ['Views', 'views'])
                df_igp_filtered.rename(columns={pub_time_col: 'Post Date', views_col: 'Impressions'}, inplace=True)

                final_igp_cols = [
                    '_Period', '_Year', '_Month', '_Day',
                    'Post Date', 'Type', 'Post Message', 'Pillar', 'Category',
                    'Sub-Category', 'Campaign Name', 'Total Post Reach',
                    'Impressions', 'Likes', 'Comments', 'Shares', 'Saves', 'Total Interactions'
                ]
                for col in final_igp_cols:
                    if col not in df_igp_filtered.columns:
                        df_igp_filtered[col] = ""

                df_igp_filtered['Total_Interactions_Numeric'] = clean_numeric_col(df_igp_filtered, 'Total Interactions')

                pivot_igp_pillar = df_igp_filtered.groupby(['_Period', 'Pillar']).agg(
                    No_of_post=('Pillar', 'count'),
                    Sum_Interactions=('Total_Interactions_Numeric', 'sum')
                ).reset_index().rename(columns={'_Period': 'Period', 'No_of_post': 'No. of Posts', 'Sum_Interactions': 'Total Interactions'})

                df_igp_filtered = df_igp_filtered.drop(columns=['Total_Interactions_Numeric'])

                df_igp_for_metrics = df_igp_filtered.copy()

                df_igp_filtered = df_igp_filtered[final_igp_cols]
                for col in ['Total Post Reach', 'Impressions', 'Likes', 'Comments', 'Shares', 'Saves', 'Total Interactions']:
                    if col in df_igp_filtered.columns:
                        df_igp_filtered[col] = pd.to_numeric(
                            df_igp_filtered[col].astype(str).str.replace(',', ''), errors='coerce')

                write_to_sheet("IG Wall Post Performance", df_igp_filtered)
                write_to_sheet("IG Pivot (Pillar)", pivot_igp_pillar)
    except Exception as e:
        logging.error(f"IG Wall Post error: {e}")

    # ------------------------------------------------------------------
    # 4. IG KEY METRICS + IG KEY METRIC (CM)
    # ------------------------------------------------------------------
    try:
        ig_post_for_metrics = df_igp_for_metrics if not df_igp_for_metrics.empty else None
        ig_story_for_metrics = df_igs_for_metrics if not df_igs_for_metrics.empty else None

        df_ig_key = generate_ig_key_metrics_all(ig_post_for_metrics, ig_story_for_metrics)
        if not df_ig_key.empty:
            df_ig_key = add_followers_to_key_metrics(df_ig_key, ig_followers_map)
            write_to_sheet("IG Key Metrics", df_ig_key)
    except Exception as e:
        logging.error(f"IG Key Metrics error: {e}")

    try:
        ig_post_for_cm = df_igp_for_metrics.copy() if not df_igp_for_metrics.empty else None
        ig_story_for_cm = df_igs_for_metrics.copy() if not df_igs_for_metrics.empty else None

        df_ig_key_cm = generate_ig_key_metrics_cm_all(ig_post_for_cm, ig_story_for_cm)
        if not df_ig_key_cm.empty:
            df_ig_key_cm = add_followers_to_key_metrics(df_ig_key_cm, ig_followers_map)
            write_to_sheet("IG Key Metric (CM)", df_ig_key_cm)
    except Exception as e:
        logging.error(f"IG Key Metric (CM) error: {e}")

    # ------------------------------------------------------------------
    # 5. COMMENTS (from ALL cleaned Excel file)
    # ------------------------------------------------------------------
    try:
        if os.path.exists(COMMENTS_XLSX):
            logging.info(f"Loading cleaned comments: {COMMENTS_XLSX.name} ...")
            df_comments = pd.read_excel(COMMENTS_XLSX)
            if 'Post Date' in df_comments.columns:
                parsed_post = smart_parse_dates(df_comments['Post Date'].astype(str).str.strip())
                df_comments['Post Date'] = parsed_post.dt.strftime('%Y-%m-%d %H:%M')
            if 'Comment Date' in df_comments.columns:
                month_hints = df_comments['Month Note'] if 'Month Note' in df_comments.columns else None
                parsed_comment_dates = smart_parse_dates(
                    df_comments['Comment Date'].astype(str).str.strip(),
                    month_hints=month_hints
                )
                df_comments['_Period'] = parsed_comment_dates.dt.strftime('%Y-%m')
                df_comments['_Year'] = parsed_comment_dates.dt.year.fillna(0).astype(int)
                df_comments['_Month'] = parsed_comment_dates.dt.month.fillna(0).astype(int)
                df_comments['_Day'] = parsed_comment_dates.dt.day.fillna(0).astype(int)
            elif 'Post Date' in df_comments.columns:
                parsed_comment_dates = parsed_post
                df_comments['_Period'] = parsed_comment_dates.dt.strftime('%Y-%m')
                df_comments['_Year'] = parsed_comment_dates.dt.year.fillna(0).astype(int)
                df_comments['_Month'] = parsed_comment_dates.dt.month.fillna(0).astype(int)
                df_comments['_Day'] = parsed_comment_dates.dt.day.fillna(0).astype(int)
            write_to_sheet("Master Comments Base", df_comments)

            if not df_comments.empty and 'Sentiment' in df_comments.columns:
                df_comments['Sentiment'] = df_comments['Sentiment'].astype(str).str.strip().str.title().replace('N/a', 'N/A')
                df_comments = df_comments[df_comments['Sentiment'].isin(['Positive', 'Neutral', 'Negative', 'N/A'])]

                if 'Category' in df_comments.columns and '_Period' in df_comments.columns:
                    sent_cat_pivot = df_comments.groupby(['_Period', 'Category', 'Sentiment']).size().unstack(fill_value=0).reset_index()
                    for s in ['Positive', 'Neutral', 'Negative', 'N/A']:
                        if s not in sent_cat_pivot.columns:
                            sent_cat_pivot[s] = 0
                    sent_cat_pivot['Total'] = sent_cat_pivot[['Positive', 'Neutral', 'Negative', 'N/A']].sum(axis=1)
                    sent_cat_pivot = sent_cat_pivot.sort_values(['_Period', 'Total'], ascending=[True, False])
                    sent_cat_pivot = sent_cat_pivot.rename(columns={'_Period': 'Period'})
                    period_col = sent_cat_pivot.pop('Period')
                    sent_cat_pivot.insert(0, 'Period', period_col)
                    write_to_sheet("Sentiment Summary (Category)", sent_cat_pivot)

                type_col = 'Type' if 'Type' in df_comments.columns else ('type' if 'type' in df_comments.columns else None)
                if type_col and '_Period' in df_comments.columns:
                    sent_type_pivot = df_comments.groupby(['_Period', type_col, 'Sentiment']).size().unstack(fill_value=0).reset_index()
                    for s in ['Positive', 'Neutral', 'Negative', 'N/A']:
                        if s not in sent_type_pivot.columns:
                            sent_type_pivot[s] = 0
                    sent_type_pivot['Total'] = sent_type_pivot[['Positive', 'Neutral', 'Negative', 'N/A']].sum(axis=1)
                    sent_type_pivot = sent_type_pivot.sort_values(['_Period', 'Total'], ascending=[True, False])
                    sent_type_pivot = sent_type_pivot.rename(columns={'_Period': 'Period'})
                    period_col = sent_type_pivot.pop('Period')
                    sent_type_pivot.insert(0, 'Period', period_col)
                    write_to_sheet("Sentiment Summary (Type)", sent_type_pivot)
        else:
            logging.warning(f"Comments Excel file not found: {COMMENTS_XLSX}")
    except Exception as e:
        logging.error(f"Comments Excel error: {e}")

    # ------------------------------------------------------------------
    # 5b. ADD API-SOURCED TABS + REACH FUNNEL
    # ------------------------------------------------------------------
    api_tabs = build_api_tabs()
    for tab_name, df_api in api_tabs.items():
        write_to_sheet(tab_name, df_api)

    df_reach_funnel = build_reach_funnel_tab()
    if df_reach_funnel is not None and not df_reach_funnel.empty:
        write_to_sheet("FB Reach Funnel", df_reach_funnel)

    # ------------------------------------------------------------------
    # 5c. PIPELINE LINKEDIN ATTACHMENTS (UNFILTERED TIMELINE PROCESSING)
    # ------------------------------------------------------------------
    logging.info("Processing LinkedIn data tabs across all parameters ...")
    try:
        write_to_sheet("LinkedIn Posts Perf", build_linkedin_posts_perf())
        write_to_sheet("LinkedIn Page Perf", build_linkedin_page_perf())
        write_to_sheet("LinkedIn Follower Log", build_linkedin_follower_log())
    except Exception as e:
        logging.error(f"LinkedIn data aggregation error: {e}")

    # ------------------------------------------------------------------
    # 5d. BACKFILL MISSING TABS
    # ------------------------------------------------------------------
    for tab in EXPECTED_TABS:
        if tab not in output_frames or output_frames[tab].empty:
            placeholder = pd.DataFrame({
                'Period': ['N/A'],
                'Note': ['No data available'],
            })
            output_frames[tab] = placeholder
            logging.info(f"  Backfilled empty tab: '{tab}'")

    # ------------------------------------------------------------------
    # 6. EXPORT TO EXCEL
    # ------------------------------------------------------------------
    logging.info(f"Exporting to Excel: {FEED_OUTPUT}")
    if output_frames:
        API_TAB_ORDER = ["Unique Page View", "FB Followers", "IG Followers", "FB Reach Funnel"]
        ordered_sheets = [t for t in EXPECTED_TABS if t in output_frames] + [t for t in API_TAB_ORDER if t in output_frames]

        logging.info(f"Exporting {len(ordered_sheets)} sheets to Excel: {FEED_OUTPUT}")
        with pd.ExcelWriter(str(FEED_OUTPUT), engine='openpyxl') as writer:
            for sheet_name in ordered_sheets:
                df_out = output_frames[sheet_name]
                if "LinkedIn" not in sheet_name:
                    df_out = normalize_date_columns(df_out)
                clean_df = deep_clean_df(df_out.fillna(""))
                safe_name = sheet_name[:31]
                clean_df.to_excel(writer, sheet_name=safe_name, index=False)

            from openpyxl.styles import Alignment

            INT_KW = ['reach', 'likes', 'comment', 'share', 'save', 'interaction',
                      'impression', 'click', 'follower', 'growth', 'no. of',
                      'posted', 'paid', 'wall post', 'view', 'reactions',
                      'sum of link', 'count', 'dark post', 'organic', 'video',
                      'feeds', 'reposts', 'follows', 'desktop', 'mobile', 'total']
            PCT_KW = ['rate', '%']
            DP1_KW = ['average']
            PCT0_KW = ['organic %', 'paid %']

            def col_fmt(header):
                h = str(header).lower()
                if any(kw in h for kw in PCT0_KW):
                    return '0%'
                if any(kw in h for kw in PCT_KW):
                    return '0.00%'
                if any(kw in h for kw in DP1_KW):
                    return '#,##0.0'
                if any(kw in h for kw in INT_KW):
                    return '#,##0'
                return None

            def row_fmt(metric_label):
                h = str(metric_label).lower()
                if any(kw in h for kw in PCT0_KW):
                    return '0%'
                if any(kw in h for kw in PCT_KW):
                    return '0.00%'
                if any(kw in h for kw in DP1_KW):
                    return '#,##0.0'
                if any(kw in h for kw in INT_KW):
                    return '#,##0'
                return None

            ROW_FMT_TABS = ['FB Key Metrics', 'FB Key Metric (CM)', 'IG Key Metrics', 'IG Key Metric (CM)']

            for ws in writer.book.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        cur = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=cur.horizontal,
                            vertical=cur.vertical,
                            wrap_text=False,
                        )

                if ws.title in ROW_FMT_TABS:
                    for row in ws.iter_rows(min_row=2, min_col=1):
                        label = row[0].value
                        fmt = row_fmt(label)
                        if fmt:
                            for cell in row[1:]:
                                cell.number_format = fmt
                else:
                    headers = {}
                    for cell in ws[1]:
                        headers[cell.column] = cell.value
                    for col_idx, header in headers.items():
                        fmt = col_fmt(header)
                        if fmt:
                            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    cell.number_format = fmt

        logging.info(f"DONE! Excel exported: {FEED_OUTPUT}")
    else:
        logging.warning("No data to export.")


if __name__ == "__main__":
    main()